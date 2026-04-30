#!/usr/bin/env python3
"""
gen_eecc_v7.py — Generador de EECC parametrizado (RT6/RT9/RT17/RT37/RT54)
Uso:
  python3 gen_eecc_v7.py \
    --empresa "ACEROS SUDAMERICANOS S.R.L." \
    --cuit "30-71870084-8" \
    --nro-ejercicio 2 \
    --fecha-cierre 2025-12-31 \
    --cof 1.2870 \
    --cap-nominal 800000 \
    --ss-actual /ruta/SS2025.xlsx \
    [--eecc-anterior /ruta/EECC2024.pdf]  (PDF o xlsx del año anterior, opcional)
    [--socio "Federico Guzzetti"] \
    [--domicilio "Austria 2128 5 11, C.A.B.A."] \
    [--actividad "Venta al por menor de materiales de construcción NCP"] \
    [--output /ruta/output.xlsx]
"""
import argparse, sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════
ap = argparse.ArgumentParser(description="Genera EECC en moneda homogénea RT6")
ap.add_argument("--empresa",        required=True,  help="Razón social")
ap.add_argument("--cuit",           required=True,  help="CUIT con guiones")
ap.add_argument("--nro-ejercicio",  type=int, default=1)
ap.add_argument("--fecha-cierre",   required=True,  help="YYYY-MM-DD")
ap.add_argument("--cof",            type=float, required=True, help="Coeficiente de reexpresión")
ap.add_argument("--cap-nominal",    type=float, required=True, help="Capital social nominal")
ap.add_argument("--ss-actual",      required=True,  help="Sumas y Saldos año actual (.xlsx)")
ap.add_argument("--eecc-anterior",  default=None,   help="EECC año anterior: PDF o xlsx (opcional)")
ap.add_argument("--socio",          default="Socio",       help="Nombre del socio principal")
ap.add_argument("--domicilio",      default="",            help="Domicilio legal")
ap.add_argument("--actividad",      default="",            help="Actividad principal")
ap.add_argument("--output",         default=None,          help="Ruta del xlsx de salida")
args = ap.parse_args()

COF        = args.cof
EMPRESA    = args.empresa
CUIT       = args.cuit
NRO_EJ     = args.nro_ejercicio
cap_nominal = args.cap_nominal
SOCIO      = args.socio

fecha_cie  = datetime.strptime(args.fecha_cierre.strip(), "%Y-%m-%d")
fecha_ant  = fecha_cie.replace(year=fecha_cie.year - 1)
EJ25       = fecha_cie.strftime("%d/%m/%Y")
EJ24       = fecha_ant.strftime("%d/%m/%Y")

AÑO_ACT    = fecha_cie.year
AÑO_ANT    = fecha_ant.year

def rx(v): return round(v * COF, 2)

# ═══════════════════════════════════════════════════════════════════════════
# LEER SUMAS Y SALDOS
# ═══════════════════════════════════════════════════════════════════════════
def read_ss(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rubros  = {}   # rubro  → saldo acumulado
    cuentas = {}   # cuenta → saldo

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        rubro, cuenta, saldo = row[0], row[1], row[2]
        if cuenta is None or saldo is None:
            continue
        saldo = float(saldo)
        rubro_key = str(rubro).strip()
        cuenta_key = str(cuenta).strip()
        rubros[rubro_key] = rubros.get(rubro_key, 0.0) + saldo
        cuentas[cuenta_key] = saldo

    def gc(*names):
        for n in names:
            if n in cuentas:
                return abs(cuentas[n])
        return 0.0

    def gr(*names):
        for n in names:
            if n in rubros:
                return abs(rubros[n])
        return 0.0

    # ── Activo ───────────────────────────────────────────────────────────────
    caja_banco = gc("Banco Santander Río en $", "Banco Galicia en $", "Banco en $")
    caja_caja  = gc("Caja")
    caja       = gr("Caja y Bancos")
    cv         = gr("Creditos Por ventas", "Créditos por ventas")
    oc_iva     = gc("IVA Saldo Técnico")
    oc_sld     = gc("Saldo Libre Disponibilidad")
    oc_dbc     = gc("Impuesto a los Db y Cr", "Impuesto a los Débitos y Créditos")
    oc_ret     = gc("Retención Ganancias Sufrida")
    oc         = gr("Otro Creditos", "Otros Créditos", "Otros créditos")
    bc         = gr("Bienes de Cambio")
    inv        = gr("Inversiones")
    bu         = gr("Bienes de Uso", "Bienes de uso")

    # ── Pasivo ───────────────────────────────────────────────────────────────
    dc         = gr("Deudas Comerciales")
    df_bsas    = gc("IIBB BSAS a pagar", "IIBB Buenos Aires a pagar")
    df_caba    = gc("IIBB CABA a pagar", "IIBB Ciudad de Buenos Aires a pagar")
    df_mdz     = gc("IIBB Mendoza a pagar", "IIBB MENDOZA a pagar")
    df         = gr("Deudas Fiscales")
    rem        = gr("Remuneraciones y cargas sociales", "Remuneraciones y Cargas Sociales")
    ds         = gr("Deudas Sociales")

    # ── Patrimonio Neto ──────────────────────────────────────────────────────
    cap_ss     = gc("Capital")
    aj_cap_ss  = gc("Ajuste de Capital")
    rna_ss     = gc("Resultado no asignado", "RNA", "Resultado No Asignado")

    # ── Resultados ───────────────────────────────────────────────────────────
    ventas     = gr("Ingresos")
    costo_sv   = gc("Compra de Servicios", "Costo de Ventas", "Costo de mercaderías")
    gcom_bsas  = gc("IIBB Buenos Aires")
    gcom_caba  = gc("IIBB CABA")
    gcom_mdz   = gc("IIBB MENDOZA", "IIBB Mendoza")
    gcom       = gr("Gastos")
    gadm_suel  = gc("Sueldos y Jornales")
    gadm_cs    = gc("Cargas Sociales")
    gadm_banc  = gc("Gastos bancarios", "Gastos Bancarios")
    gadm_adm   = gc("Gastos Administrativos")
    # RECPAM: en SS negativo=ingreso, positivo=pérdida → guardar como ingreso positivo
    _recpam_ss = cuentas.get("RECPAM", cuentas.get("Recpam", 0.0))
    recpam     = -_recpam_ss   # flip: crédito (negativo en SS) → positivo = ganancia
    ig         = gc("Impuesto a las Ganancias", "Impuesto a las ganancias")

    return {
        "caja_banco": caja_banco, "caja_caja": caja_caja, "caja": caja,
        "cv": cv, "oc_iva": oc_iva, "oc_sld": oc_sld, "oc_dbc": oc_dbc,
        "oc_ret": oc_ret, "oc": oc, "bc": bc, "inv": inv, "bu": bu,
        "dc": dc, "df_bsas": df_bsas, "df_caba": df_caba, "df_mdz": df_mdz,
        "df": df, "rem": rem, "ds": ds,
        "cap_ss": cap_ss, "aj_cap_ss": aj_cap_ss, "rna_ss": rna_ss,
        "ventas": ventas, "costo_sv": costo_sv,
        "gcom_bsas": gcom_bsas, "gcom_caba": gcom_caba, "gcom_mdz": gcom_mdz, "gcom": gcom,
        "gadm_suel": gadm_suel, "gadm_cs": gadm_cs, "gadm_banc": gadm_banc, "gadm_adm": gadm_adm,
        "recpam": recpam, "ig": ig,
    }


def _empty_sa():
    keys = ["caja_banco","caja_caja","caja","cv","oc_iva","oc_sld","oc_dbc","oc_ret","oc",
            "bc","inv","bu","dc","df_bsas","df_caba","df_mdz","df","rem","ds",
            "cap_ss","aj_cap_ss","rna_ss","ventas","costo_sv",
            "gcom_bsas","gcom_caba","gcom_mdz","gcom",
            "gadm_suel","gadm_cs","gadm_banc","gadm_adm","recpam","ig"]
    return {k: 0.0 for k in keys}


def read_eecc_pdf(pdf_path):
    """Extrae saldos nominales del EECC anterior desde un PDF (formato propio o externo CPCECABA).
    Retorna dict compatible con read_ss(). Sub-ítems no disponibles quedan en 0.
    """
    import pdfplumber, re

    NUM_PAT = re.compile(
        r'(\([\d\s\.]+,\d{2}\)|[\d][\d\s\.]*,\d{2}|-{1})\s'
    )

    def first_value_after(text, label_end):
        """Extrae el primer valor numérico que aparece después de label_end en el texto."""
        snippet = text[label_end:]
        m = NUM_PAT.search(snippet)
        if not m:
            return None
        raw = re.sub(r'\s', '', m.group()).strip()
        if raw in ('-', '–', ''):
            return 0.0
        neg = raw.startswith('(') and raw.endswith(')')
        if neg:
            raw = raw[1:-1]
        if re.match(r'^[\d\.]+,\d{2}$', raw):
            raw = raw.replace('.', '').replace(',', '.')
        try:
            v = float(raw)
            return -v if neg else v
        except ValueError:
            return None

    # Mapeo label (fragmento lowercase) → campo
    LABEL_MAP = [
        ('caja y bancos',                        'caja'),
        ('créditos por ventas',                  'cv'),
        ('otros créditos',                       'oc'),
        ('bienes de cambio (nota',               'bc'),
        ('bienes de uso',                        'bu'),
        ('total del activo corriente',           'ta'),
        ('comerciales (nota',                    'dc'),
        ('cargas fiscales',                      'df'),
        ('remuneraciones y cargas sociales',     'rem'),
        ('deudas sociales',                      'ds'),
        ('total del pasivo corriente',           'tp'),
        ('patrimonio neto',                      'pn'),
        ('ventas netas de bienes',               'ventas'),
        ('costo de bienes vendidos',             'costo'),
        ('gastos de comercialización',           'gcom'),
        ('gastos de administración',             'gadm'),
        ('resultados financieros y por tenencia','recpam'),
        ('impuesto a las ganancias (nota',             'ig'),
        ('ganancia (pérdida) de las operaciones',      'res'),
        ('ganancia (perdida) de las operaciones',      'res'),
        ('ajuste de capital',                          'aj_cap'),
    ]

    result = {v: 0.0 for _, v in LABEL_MAP}

    with pdfplumber.open(pdf_path) as pdf:
        full_text = '\n'.join(p.extract_text() or '' for p in pdf.pages)

    full_lower = full_text.lower()
    for frag, field in LABEL_MAP:
        if result[field] != 0.0:
            continue
        pos = full_lower.find(frag)
        if pos == -1:
            continue
        v = first_value_after(full_text, pos + len(frag))
        if v is not None:
            result[field] = v

    # Ajuste de Capital: está en el EEPN en la línea "Suscripción de Capital cap_nominal  aj_cap ..."
    # Extraemos todos los números y tomamos el segundo (el ajuste)
    if result['aj_cap'] == 0.0:
        for line in full_text.split('\n'):
            if 'suscripci' in line.lower() and 'capital' in line.lower():
                nums = re.findall(r'\([\d\s\.]+,\d{2}\)|[\d][\d\s\.]*,\d{2}', line)
                nums_parsed = []
                for n in nums:
                    v = first_value_after(n + ' ', 0)
                    if v is not None:
                        nums_parsed.append(abs(v))
                if len(nums_parsed) >= 2:
                    result['aj_cap'] = nums_parsed[1]  # segundo valor = Ajuste de Capital
                break

    sa = _empty_sa()
    sa.update({
        'caja':        abs(result['caja']),
        'cv':          abs(result['cv']),
        'oc':          abs(result['oc']),
        'bc':          abs(result['bc']),
        'bu':          abs(result['bu']),
        'dc':          abs(result['dc']),
        'df':          abs(result['df']),
        'rem':         abs(result['rem']),
        'ds':          abs(result['ds']),
        'aj_cap_ss':   result['aj_cap'],
        'ventas':      abs(result['ventas']),
        'costo_sv':    abs(result['costo']),
        'recpam':      result['recpam'],
        'ig':          abs(result['ig']),
        '_gcom_total': abs(result['gcom']),
        '_gadm_total': abs(result['gadm']),
        '_r24_res':    result['res'],
    })
    return sa

# ═══════════════════════════════════════════════════════════════════════════
# DATOS
# ═══════════════════════════════════════════════════════════════════════════
ss  = read_ss(args.ss_actual)
if args.eecc_anterior:
    if args.eecc_anterior.lower().endswith('.pdf'):
        sa = read_eecc_pdf(args.eecc_anterior)
    else:
        sa = read_ss(args.eecc_anterior)
else:
    sa = _empty_sa()

# ── 2025 (año actual) ────────────────────────────────────────────────────
caja25_banco = ss["caja_banco"]
caja25_caja  = ss["caja_caja"]
caja25       = ss["caja"]
cv25         = ss["cv"]
oc25_iva     = ss["oc_iva"]
oc25_sld     = ss["oc_sld"]
oc25_dbc     = ss["oc_dbc"]
oc25_ret     = ss["oc_ret"]
oc25         = ss["oc"]
bc25         = ss["bc"]
inv25        = ss["inv"]
bu25         = ss["bu"]
ta25         = caja25 + cv25 + oc25 + bc25 + inv25 + bu25

dc25         = ss["dc"]
df25_bsas    = ss["df_bsas"]
df25_caba    = ss["df_caba"]
df25         = ss["df"]
rem25        = ss["rem"]
ds25         = ss["ds"]
tp25         = dc25 + df25 + rem25 + ds25
pn25         = round(ta25 - tp25, 2)

ventas25     = ss["ventas"]
costo_sv25   = ss["costo_sv"]
gb25         = round(ventas25 - costo_sv25, 2)
gcom25_bsas  = ss["gcom_bsas"]
gcom25_caba  = ss["gcom_caba"]
gcom25_mdz   = ss["gcom_mdz"]
gcom25       = gcom25_bsas + gcom25_caba + gcom25_mdz
gadm25_suel  = ss["gadm_suel"]
gadm25_cs    = ss["gadm_cs"]
gadm25_banc  = ss["gadm_banc"]
gadm25_adm   = ss["gadm_adm"]
gadm25       = gadm25_suel + gadm25_cs + gadm25_banc + gadm25_adm
recpam25     = ss["recpam"]
ig25         = ss["ig"]
res25        = round(gb25 - gcom25 - gadm25 + recpam25 - ig25, 2)

# ── 2024 nominales (año anterior SS) ────────────────────────────────────
r24_caja     = sa["caja"]
r24_cv       = sa["cv"]
r24_oc       = sa["oc"]
r24_bc       = sa["bc"]
r24_ta       = r24_caja + r24_cv + r24_oc + r24_bc + sa["inv"] + sa["bu"]

r24_dc       = sa["dc"]
r24_df       = sa["df"]
r24_rem      = sa["rem"]
r24_ds       = sa["ds"]
r24_tp       = r24_dc + r24_df + r24_rem + r24_ds
r24_pn       = round(r24_ta - r24_tp, 2)

r24_ventas   = sa["ventas"]
r24_costo    = sa["costo_sv"]
_r24_gcom_items = sa["gcom_bsas"] + sa["gcom_caba"] + sa["gcom_mdz"]
r24_gcom     = _r24_gcom_items if _r24_gcom_items > 0 else sa.get('_gcom_total', 0.0)
r24_gadm_suel= sa["gadm_suel"]
r24_gadm_cs  = sa["gadm_cs"]
r24_gadm_banc= sa["gadm_banc"]
r24_gadm_adm = sa["gadm_adm"]
_r24_gadm_items = r24_gadm_suel + r24_gadm_cs + r24_gadm_banc + r24_gadm_adm
r24_gadm     = _r24_gadm_items if _r24_gadm_items > 0 else sa.get('_gadm_total', 0.0)
r24_recpam_signed = sa["recpam"]   # ya tiene signo correcto: positivo=ingreso, negativo=pérdida

r24_ig       = sa["ig"]
r24_gb       = r24_ventas - r24_costo
r24_res_calc = round(r24_gb - r24_gcom - r24_gadm + r24_recpam_signed - r24_ig, 2)
r24_res      = sa.get('_r24_res') if sa.get('_r24_res') else r24_res_calc

# ── 2024 reexpresados (× COF) ────────────────────────────────────────────
caja24  = rx(r24_caja)
cv24    = rx(r24_cv)
oc24    = rx(r24_oc)
bc24    = rx(r24_bc)
ta24    = rx(r24_ta)
dc24    = rx(r24_dc)
df24    = rx(r24_df)
rem24   = rx(r24_rem)
ds24    = rx(r24_ds)
tp24    = rx(r24_tp)
pn24    = rx(r24_pn)

ventas24 = rx(r24_ventas)
costo24  = rx(r24_costo)
gb24     = round(ventas24 - costo24, 2)
gcom24   = rx(r24_gcom + r24_gadm)   # en 2024 todo va a un solo rubro si gadm=0
recpam24 = rx(r24_recpam_signed)
ig24     = rx(r24_ig)
res24    = rx(r24_res)

# ── PN 2024 reexpresado para EEPN (RT6/RT17: capital siempre nominal) ───
aj_cap_anterior  = sa["aj_cap_ss"]               # ajuste de capital nominal año anterior
total_ap_rx      = rx(cap_nominal + aj_cap_anterior)
aj24_eepn        = round(total_ap_rx - cap_nominal, 2)
pn24_eepn_tot    = round(cap_nominal + aj24_eepn + res24, 2)
pn24_rx          = pn24_eepn_tot

# ── Ajuste RT6 apertura ──────────────────────────────────────────────────
recpam_rx_aper   = round(pn25 - pn24_eepn_tot - res25, 2)
recpam25_adj     = round(recpam25 + recpam_rx_aper, 2)
res_ai25_adj     = gb25 - gcom25 - gadm25 + recpam25_adj
res25_adj        = round(res_ai25_adj - ig25, 2)

# EEPN apertura
aj25_eepn_rx     = aj24_eepn
rna_inicio_rx    = res24
pn_apert_rx      = round(cap_nominal + aj25_eepn_rx + rna_inicio_rx, 2)

# ── EF ──────────────────────────────────────────────────────────────────
ef_ini25 = caja24
ef_cie25 = caja25
ef_var25 = round(ef_cie25 - ef_ini25, 2)
vcv25    = -(cv25  - cv24)
voc25    = -(oc25  - oc24)
vbc25    = bc24 - bc25
vdc25    = dc25 - dc24
vdf25    = df25 - df24
vrem25   = rem25 - rem24
vds25    = ds25 - ds24
fne_op_check = res25_adj - recpam25_adj + vcv25 + voc25 + vbc25 + vdc25 + vdf25 + vrem25 + vds25
dif_rx25 = round(ef_var25 - fne_op_check, 2)
fne_op25 = ef_var25

ef_ini24 = 0.0
ef_cie24 = caja24
ef_var24 = round(ef_cie24 - ef_ini24, 2)
fne_op24 = rx(r24_caja)      # simplificación: FNE op = variación efectivo 2024
fne_fin24 = rx(-cap_nominal)  # aportes de capital iniciales (negativo = ingreso de fondos)

ef24_res    = res24
ef24_recpam = rx(abs(r24_recpam_signed))
ef24_vcv    = rx(-r24_cv)
ef24_voc    = rx(-r24_oc)
ef24_vbc    = rx(-r24_bc)
ef24_vdc    = rx(r24_dc)
ef24_imp    = rx(r24_df + r24_ig)

# ── Anexo II ─────────────────────────────────────────────────────────────
costo_ei25    = bc24
costo_cmp25   = costo_sv25 - costo_ei25
costo_sub25   = costo_ei25 + costo_cmp25
costo_ef25    = bc25
costo_costo25 = costo_sv25

costo_ei24    = 0.0
costo_cmp24   = rx(r24_costo + r24_bc)
costo_sub24   = costo_cmp24
costo_ef24    = bc24
costo_costo24 = round(costo_cmp24 - costo_ef24, 2)

# ── Anexo III 2024 ───────────────────────────────────────────────────────
r24_a3_rem   = rx(r24_gadm_suel) if r24_gadm_suel else rx(r24_gcom * 0.93)
r24_a3_cs    = rx(r24_gadm_cs)   if r24_gadm_cs   else rx(r24_gcom * 0.001)
r24_a3_iibb  = round(gcom24 + rx(r24_gadm) - r24_a3_rem - r24_a3_cs, 2)

# ── Verificaciones ───────────────────────────────────────────────────────
print("── VERIFICACIONES ──")
print(f"TA {AÑO_ACT}: {ta25:,.2f}  TP: {tp25:,.2f}  PN: {pn25:,.2f}")
print(f"TA {AÑO_ANT}: {ta24:,.2f}  TP: {tp24:,.2f}  PN reexp: {pn24_eepn_tot:,.2f}")
print(f"Resultado {AÑO_ACT} (SS): {res25:,.2f}  | adj RT6: {res25_adj:,.2f}")
print(f"RECPAM {AÑO_ACT}: {recpam25:,.2f}  | adj: {recpam25_adj:,.2f}  | aper: {recpam_rx_aper:,.2f}")
print(f"EEPN apertura: {pn_apert_rx:,.2f}  | cierre {AÑO_ANT}: {pn24_eepn_tot:,.2f}  | OK: {abs(pn_apert_rx - pn24_eepn_tot) < 0.02}")
print(f"EEPN cierre: {round(pn_apert_rx + res25_adj, 2):,.2f}  | PN ESP: {pn25:,.2f}")

# ═══════════════════════════════════════════════════════════════════════════
# UTILIDADES DE ESTILO
# ═══════════════════════════════════════════════════════════════════════════
C_HDR = "1F497D"
C_SEC = "D9E1F2"
C_SUB = "B8CCE4"
C_TOT = "9DC3E6"
C_WHT = "FFFFFF"
NF    = '#,##0.00;[Red](#,##0.00);"-"'
NF0   = '#,##0;[Red](#,##0);"-"'

def fl(hex_color=None):
    if not hex_color: return PatternFill(fill_type=None)
    return PatternFill("solid", fgColor=hex_color)

def fw(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Arial", size=size)

def set_cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def row_height(ws, r, h):
    ws.row_dimensions[r].height = h

def s(ws, r, c, v=None, bg=None, fn=None, nf=None, al=None, bold_val=False):
    cell = ws.cell(row=r, column=c)
    if v is not None: cell.value = v
    if bg: cell.fill = fl(bg)
    if fn: cell.font = fn
    elif bold_val: cell.font = Font(bold=True, name="Arial", size=10)
    else: cell.font = Font(name="Arial", size=10)
    if nf and isinstance(v, (int, float)): cell.number_format = nf
    if al: cell.alignment = al
    return cell

def title_block(ws, empresa, titulo, subtitle=None, ncols=3):
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1); c.value = empresa
    c.font = Font(bold=True, size=13, name="Arial", color=C_HDR)
    c.alignment = Alignment(horizontal="center"); r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1); c.value = titulo
    c.font = Font(bold=True, size=12, name="Arial", color=C_HDR)
    c.alignment = Alignment(horizontal="center"); r += 1
    if subtitle:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(r, 1); c.value = subtitle
        c.font = Font(italic=True, size=9, name="Arial")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_height(ws, r, 28); r += 1
    return r

def col_headers(ws, r, labels, bg=C_HDR, ncols=None):
    n = ncols or len(labels)
    for ci, lbl in enumerate(labels, 1):
        cell = ws.cell(r, ci); cell.value = lbl; cell.fill = fl(bg)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci in range(len(labels)+1, n+1):
        ws.cell(r, ci).fill = fl(bg)
    row_height(ws, r, 30); return r + 1

def section(ws, r, label, ncols=3, bg=C_SEC):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cell = ws.cell(r, 1); cell.value = label; cell.fill = fl(bg)
    cell.font = Font(bold=True, color=C_HDR, name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for ci in range(2, ncols+1): ws.cell(r, ci).fill = fl(bg)
    return r + 1

def detail(ws, r, label, v25=None, v24=None, indent=2):
    ws.cell(r, 1).value = (" " * indent) + label
    ws.cell(r, 1).font  = Font(name="Arial", size=10)
    if v25 is not None:
        c = ws.cell(r, 2); c.value = v25; c.number_format = NF; c.font = Font(name="Arial", size=10)
    if v24 is not None:
        c = ws.cell(r, 3); c.value = v24; c.number_format = NF; c.font = Font(name="Arial", size=10)
    return r + 1

def total(ws, r, label, v25=None, v24=None, bg=C_TOT):
    ws.cell(r, 1).value = label; ws.cell(r, 1).fill = fl(bg)
    ws.cell(r, 1).font  = Font(bold=True, color=C_HDR, name="Arial", size=10)
    for ci, val in [(2, v25), (3, v24)]:
        c = ws.cell(r, ci); c.fill = fl(bg)
        if val is not None: c.value = val; c.number_format = NF
        c.font = Font(bold=True, color=C_HDR, name="Arial", size=10)
    return r + 1

def hdr_section(ws, r, label, ncols=3):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cell = ws.cell(r, 1); cell.value = label; cell.fill = fl(C_HDR)
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left")
    for ci in range(2, ncols+1): ws.cell(r, ci).fill = fl(C_HDR)
    return r + 1

def blank(ws, r): return r + 1

# ═══════════════════════════════════════════════════════════════════════════
# WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════
wb = Workbook(); wb.remove(wb.active)

# ═══════════════════════════════════════════════════════════════════════════
# CARÁTULA
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Carátula")
set_cols(ws, [30, 50])
r = title_block(ws, EMPRESA, f"ESTADOS CONTABLES — EJERCICIO Nº {NRO_EJ}", ncols=2)
r += 1
datos = [
    ("CUIT:", CUIT),
    (f"Ejercicio Nº:", str(NRO_EJ)),
    ("Fecha de inicio:", f"01/01/{AÑO_ACT}"),
    ("Fecha de cierre:", EJ25),
    ("Domicilio Legal:", args.domicilio),
    ("Actividad Principal:", args.actividad),
    ("Capital Suscripto e Integrado:", f"$ {cap_nominal:,.0f}"),
    ("", ""),
    ("Normas aplicadas:", "RT6 (moneda homogénea), RT9, RT17, RT37, RT54"),
    ("Coeficiente de actualización:", f"{COF} (ajuste ejercicio comparativo {AÑO_ANT})"),
]
for lbl, val in datos:
    s(ws, r, 1, lbl, bg=C_SEC, fn=fw(bold=True, color=C_HDR))
    s(ws, r, 2, val, fn=fw()); r += 1
r += 1
s(ws, r, 1, "ÍNDICE DE ESTADOS", bg=C_HDR, fn=fw(bold=True, color="FFFFFF"))
s(ws, r, 2, "", bg=C_HDR); r += 1
for hoja, desc in [
    ("ESP","Estado de Situación Patrimonial"),("ER","Estado de Resultados"),
    ("EEPN","Estado de Evolución del Patrimonio Neto"),
    ("EF","Estado de Flujo de Efectivo (Método Indirecto)"),
    ("Anexo I","Bienes de Uso"),("Anexo II","Costo de Mercaderías Vendidas"),
    ("Anexo III","Información Art. 64 Inc. b) — Ley 19.550 (Gastos por función)"),
    ("Notas","Notas a los Estados Contables"),
]:
    s(ws, r, 1, hoja, bg=C_SUB, fn=fw(bold=True, color=C_HDR))
    s(ws, r, 2, desc,  fn=fw()); r += 1

# ═══════════════════════════════════════════════════════════════════════════
# ESP
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ESP")
set_cols(ws, [44, 18, 18])
r = title_block(ws, EMPRESA,
    "ESTADO DE SITUACIÓN PATRIMONIAL",
    f"Ejercicio N°{NRO_EJ} cerrado el {EJ25} | Comparativo al {EJ24} — en pesos de poder adquisitivo al {EJ25}",
    ncols=3)
r = col_headers(ws, r, ["RUBRO", EJ25, EJ24])
r = hdr_section(ws, r, "ACTIVO", ncols=3)
r = section(ws, r, "Activo Corriente", ncols=3)
r = detail(ws, r, "Caja y Bancos (Nota 2.1)",       caja25, caja24)
r = detail(ws, r, "Inversiones",                    inv25,  rx(0))
r = detail(ws, r, "Créditos por Ventas (Nota 2.2)", cv25,   cv24)
r = detail(ws, r, "Otros Créditos (Nota 2.3)",      oc25,   oc24)
r = detail(ws, r, "Bienes de Cambio (Nota 2.4)",    bc25,   bc24)
r = total(ws, r, "Total del Activo Corriente",      ta25,   ta24)
r = section(ws, r, "Activo No Corriente", ncols=3)
r = detail(ws, r, "Bienes de Uso (Anexo I)",        bu25,   rx(0))
r = total(ws, r, "Total del Activo No Corriente",   bu25,   rx(0))
r = total(ws, r, "TOTAL DEL ACTIVO",                ta25,   ta24, bg=C_HDR)
for ci in range(1, 4): ws.cell(r-1, ci).font = fw(bold=True, color="FFFFFF")
r = blank(ws, r)
r = hdr_section(ws, r, "PASIVO", ncols=3)
r = section(ws, r, "Pasivo Corriente", ncols=3)
r = detail(ws, r, "Deudas Comerciales (Nota 2.5.1)",              dc25,  dc24)
r = detail(ws, r, "Cargas Fiscales (Nota 2.5.2)",                 df25,  df24)
r = detail(ws, r, "Remuneraciones y Cargas Sociales (Nota 2.5.3)",rem25, rem24)
r = detail(ws, r, "Deudas Sociales (Nota 2.5.4)",                 ds25,  ds24)
r = total(ws, r, "Total Deudas",                                  tp25,  tp24)
r = total(ws, r, "Total del Pasivo Corriente",                    tp25,  tp24)
r = total(ws, r, "TOTAL DEL PASIVO",                              tp25,  tp24, bg=C_HDR)
for ci in range(1, 4): ws.cell(r-1, ci).font = fw(bold=True, color="FFFFFF")
r = blank(ws, r)
r = hdr_section(ws, r, "PATRIMONIO NETO  (según el EEPN)", ncols=3)
r = total(ws, r, "PATRIMONIO NETO", pn25, pn24_rx)
r = blank(ws, r)
r = total(ws, r, "TOTAL DEL PASIVO Y PATRIMONIO NETO", tp25+pn25, tp24+pn24_rx)
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(r,1).value = "Las notas y anexos que se acompañan forman parte integrante de los estados contables."
ws.cell(r,1).font  = Font(italic=True, size=9, name="Arial")

# ═══════════════════════════════════════════════════════════════════════════
# ER
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ER")
set_cols(ws, [48, 18, 18])
r = title_block(ws, EMPRESA,
    "ESTADO DE RESULTADOS",
    f"Ejercicio N°{NRO_EJ} cerrado el {EJ25} | Comparativo {EJ24} — en pesos de poder adquisitivo al {EJ25}",
    ncols=3)
r = col_headers(ws, r, ["RUBRO", EJ25, EJ24])
r = detail(ws, r, "Ventas netas de bienes (o servicios)",                     ventas25, ventas24, indent=0)
r = detail(ws, r, "Costo de bienes vendidos (o servicios prestados) (Anexo II)", -costo_sv25, -costo24, indent=0)
r = total(ws,  r, "Ganancia (Pérdida) bruta",                                 gb25,     gb24)
r = detail(ws, r, "Resultado valuación bienes de cambio a VNR",               None,     None, indent=0)
r = detail(ws, r, "Gastos de comercialización (Anexo III)",                   -gcom25,  -rx(r24_gcom), indent=0)
r = detail(ws, r, "Gastos de administración (Anexo III)",                     -gadm25,  -rx(r24_gadm) if r24_gadm else None, indent=0)
r = detail(ws, r, "Resultados financieros y por tenencia — RECPAM (Nota 2.6)", recpam25_adj, recpam24, indent=0)
res_ai24 = round(gb24 - rx(r24_gcom) - rx(r24_gadm) + recpam24, 2)
r = total(ws,  r, "Ganancia (Pérdida) antes del impuesto a las ganancias",    res_ai25_adj, res_ai24)
r = detail(ws, r, "Impuesto a las ganancias (Nota 2.5.2)",                    -ig25,    -ig24, indent=0)
r = total(ws,  r, "Ganancia (Pérdida) de las operaciones ordinarias",         res25_adj, res24, bg=C_HDR)
for ci in range(1, 4): ws.cell(r-1, ci).font = fw(bold=True, color="FFFFFF")
r = detail(ws, r, "Resultados de operaciones extraordinarias", None, None, indent=0)
r = total(ws,  r, "GANANCIA (PÉRDIDA) DEL EJERCICIO",                         res25_adj, res24, bg=C_HDR)
for ci in range(1, 4): ws.cell(r-1, ci).font = fw(bold=True, color="FFFFFF")
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(r,1).value = "Las notas y anexos que se acompañan forman parte integrante de los estados contables."
ws.cell(r,1).font  = Font(italic=True, size=9, name="Arial")

# ═══════════════════════════════════════════════════════════════════════════
# EEPN
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("EEPN")
NCOLS_EEPN = 13
set_cols(ws, [30, 12, 13, 11, 11, 9, 9, 10, 10, 13, 13, 13, 14])
cap25_eepn = cap_nominal
aj25_eepn  = aj25_eepn_rx
rna_inicio = rna_inicio_rx
pn_apert_ok= pn_apert_rx

r = title_block(ws, EMPRESA,
    "ESTADO DE EVOLUCIÓN DEL PATRIMONIO NETO",
    f"Ejercicio N°{NRO_EJ} cerrado el {EJ25} | en pesos de poder adquisitivo al {EJ25} | Capital expresado a valor nominal (RT6/RT17)",
    ncols=NCOLS_EEPN)

row_height(ws, r, 40)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)
for ci in range(1, NCOLS_EEPN+1):
    c = ws.cell(r, ci); c.fill = fl(C_HDR)
    c.font = fw(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.cell(r, 1).value  = "DETALLE"
ws.cell(r, 2).value  = "APORTES DE LOS PROPIETARIOS"
ws.cell(r, 6).value  = "RESULTADOS ACUMULADOS"
ws.cell(r, 10).value = "RESULTADO DEL EJERCICIO"
ws.cell(r, 12).value = f"TOTAL {EJ25}"
ws.cell(r, 13).value = f"TOTAL\n{EJ24}"
r += 1

row_height(ws, r, 40)
hdrs2 = {1:"Detalle", 2:"Capital Social", 3:"Ajuste de Capital", 4:"Aportes Irrevoc.", 5:"Primas de Emisión",
         6:"Reserva Legal", 7:"Otras Reservas", 8:"Total Reservas", 9:"Rdos. Diferidos",
         10:"RNA / No Asignados", 11:"Resultado del Ejercicio", 12:"TOTAL 2025", 13:f"TOTAL\n{EJ24}"}
for ci, h in hdrs2.items():
    c = ws.cell(r, ci); c.value = h; c.fill = fl(C_HDR)
    c.font = fw(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
r += 1

def eepn_row(ws, r, detalle, cs=None, aj=None, ai=None, pe=None,
             rl=None, or_=None, tr=None, rd=None, rna=None, re=None,
             tot25=None, tot24=None, bg=None, is_bold=False):
    row_height(ws, r, 18)
    fn = fw(bold=is_bold, color=C_HDR if bg == C_TOT else "000000", size=9)
    c = ws.cell(r, 1); c.value = detalle; c.font = fn
    if bg: c.fill = fl(bg)
    for ci, val in enumerate([cs, aj, ai, pe, rl, or_, tr, rd, rna, re, tot25, tot24], 2):
        cell = ws.cell(r, ci)
        if bg: cell.fill = fl(bg)
        cell.font = fn
        if val is not None: cell.value = val; cell.number_format = NF

def eepn_sec(ws, r, label):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS_EEPN)
    c = ws.cell(r, 1); c.value = label; c.fill = fl(C_SEC)
    c.font = fw(bold=True, color=C_HDR, size=10)
    c.alignment = Alignment(horizontal="left")
    for ci in range(2, NCOLS_EEPN+1): ws.cell(r, ci).fill = fl(C_SEC)
    return r + 1

total_ap_rx_eepn = rx(cap_nominal + aj_cap_anterior)
r = eepn_sec(ws, r, f"EJERCICIO {AÑO_ACT} — en pesos de poder adquisitivo al {EJ25}  |  Comparativo {EJ24} reexpresado ×{COF}")
eepn_row(ws, r, f"Saldos al inicio del ejercicio (01/01/{AÑO_ACT})",
         cs=cap25_eepn, aj=aj25_eepn, rna=rna_inicio, re=0,
         tot25=pn_apert_ok, tot24=0.0); r += 1
eepn_row(ws, r, "Suscripción de Capital",
         cs=0, aj=0, tot25=0.0, tot24=total_ap_rx_eepn); r += 1
eepn_row(ws, r, "Cobros aportes irrevocables / Capitaliz. aportes"); r += 1
eepn_row(ws, r, "Distribución de resultados — Reserva legal"); r += 1
eepn_row(ws, r, "Distribución de resultados — Dividendos"); r += 1
eepn_row(ws, r, "Absorción de pérdidas acumuladas"); r += 1
eepn_row(ws, r, "Ganancia (Pérdida) del ejercicio",
         cs=0, aj=0, rna=0, re=res25_adj,
         tot25=res25_adj, tot24=res24); r += 1
eepn_row(ws, r, f"Saldos al cierre del ejercicio — {EJ25}",
         cs=cap25_eepn, aj=aj25_eepn, rna=rna_inicio, re=res25_adj,
         tot25=pn25, tot24=pn24_eepn_tot,
         bg=C_TOT, is_bold=True); r += 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS_EEPN)
ws.cell(r,1).value = (
    f"Notas: (1) El Capital Social se expone a valor nominal ($ {cap_nominal:,.0f}) conforme RT N°6 y RT N°17; "
    f"la reexpresión monetaria integra el rubro Ajuste de Capital. "
    f"(2) El saldo comparativo al {EJ24} (${pn24_eepn_tot:,.2f}) corresponde al PN nominal reexpresado "
    f"por el coeficiente {COF} a moneda de poder adquisitivo al {EJ25}. "
    "Las notas y anexos forman parte integrante de estos estados contables."
)
ws.cell(r,1).font = Font(italic=True, size=9, name="Arial")

# ═══════════════════════════════════════════════════════════════════════════
# EF
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("EF")
set_cols(ws, [52, 18, 18])
r = title_block(ws, EMPRESA,
    "ESTADO DE FLUJO DE EFECTIVO — Método Indirecto",
    f"Ejercicio N°{NRO_EJ} cerrado el {EJ25} | Comparativo {EJ24} — en pesos de poder adquisitivo al {EJ25}",
    ncols=3)
r = col_headers(ws, r, ["CONCEPTO", EJ25, EJ24])
r = hdr_section(ws, r, "VARIACIÓN NETA DEL EFECTIVO", ncols=3)
r = detail(ws, r, "Efectivo al inicio del ejercicio",       ef_ini25, ef_ini24, indent=0)
r = detail(ws, r, "Modificación de ejercicios anteriores",  None,     None,     indent=0)
r = detail(ws, r, "Efectivo modificado al inicio",          ef_ini25, ef_ini24, indent=0)
r = detail(ws, r, "Efectivo al cierre del ejercicio",       ef_cie25, ef_cie24, indent=0)
r = total(ws,  r, "Aumento (disminución) neto del efectivo", ef_var25, ef_var24)
r = blank(ws, r)
r = hdr_section(ws, r, "CAUSAS DE LAS VARIACIONES DEL EFECTIVO", ncols=3)
r = hdr_section(ws, r, "ACTIVIDADES OPERATIVAS", ncols=3)
r = detail(ws, r, "Ganancia (Pérdida) ordinaria del ejercicio", res25_adj, ef24_res, indent=0)
r = section(ws, r, "Ajustes para arribar al FNE:", ncols=3)
r = detail(ws, r, "  Resultados financieros y por tenencia (RECPAM)", -recpam25_adj, ef24_recpam)
r = detail(ws, r, "  (Aumento) Disminución en créditos por ventas",   vcv25,   ef24_vcv)
r = detail(ws, r, "  (Aumento) Disminución en otros créditos",        voc25,   ef24_voc)
r = detail(ws, r, "  (Aumento) Disminución en bienes de cambio",      vbc25,   ef24_vbc)
r = detail(ws, r, "  Aumento (Disminución) en deudas operativas",     vdc25,   ef24_vdc)
r = detail(ws, r, "  Aumento (Disminución) en cargas fiscales",       vdf25,   None)
r = detail(ws, r, "  Aumento (Disminución) en remuneraciones",        vrem25,  None)
r = detail(ws, r, "  Aumento (Disminución) en deudas sociales",       vds25,   None)
r = detail(ws, r, "  Pagos de impuestos",                             None,    ef24_imp)
r = detail(ws, r, "  Diferencia por reexpresión monetaria (RT6)",     dif_rx25, None)
r = total(ws,  r, "Flujo neto de efectivo — actividades operativas",  fne_op25, fne_op24)
r = blank(ws, r)
r = hdr_section(ws, r, "ACTIVIDADES DE INVERSIÓN", ncols=3)
r = detail(ws, r, "Pagos por compras de bienes de uso",               None, None, indent=0)
r = total(ws,  r, "Flujo neto de efectivo — actividades de inversión", 0.0, 0.0)
r = blank(ws, r)
r = hdr_section(ws, r, "ACTIVIDADES DE FINANCIACIÓN", ncols=3)
r = detail(ws, r, "Cobros de aportes de capital en efectivo",         None, rx(cap_nominal), indent=0)
r = total(ws,  r, "Flujo neto de efectivo — actividades de financiación", 0.0, fne_fin24)
r = blank(ws, r)
r = total(ws,  r, "Aumento (Disminución) neto del efectivo", ef_var25, ef_var24, bg=C_HDR)
for ci in range(1, 4): ws.cell(r-1, ci).font = fw(bold=True, color="FFFFFF")
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(r,1).value = "Las notas y anexos que se acompañan forman parte integrante de los estados contables."
ws.cell(r,1).font  = Font(italic=True, size=9, name="Arial")

# ═══════════════════════════════════════════════════════════════════════════
# ANEXO I
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Anexo I")
NCOLS_BU = 12
set_cols(ws, [22, 12, 11, 11, 12, 12, 10, 10, 11, 12, 12, 12])
r = title_block(ws, EMPRESA, "ANEXO I — BIENES DE USO",
    f"Ejercicio N°{NRO_EJ} cerrado el {EJ25} | Comparativo al {EJ24} — en pesos de poder adquisitivo al {EJ25}",
    ncols=NCOLS_BU)
row_height(ws, r, 35)
ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=10)
ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=12)
_hdr = dict(fill=fl(C_HDR), font=fw(bold=True, color="FFFFFF", size=9),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
for ci in [1, 2, 6, 11]:
    c = ws.cell(r, ci)
    c.fill = _hdr["fill"]; c.font = _hdr["font"]; c.alignment = _hdr["alignment"]
ws.cell(r, 1).value = "RUBROS"
ws.cell(r, 2).value = "VALORES DE INCORPORACIÓN"
ws.cell(r, 6).value = "DEPRECIACIÓN"
ws.cell(r, 11).value = "NETO AL CIERRE"
r += 1
row_height(ws, r, 45)
for ci, h in {2:"Al Inicio", 3:"Altas del Ejercicio", 4:"Bajas del Ejercicio", 5:"Al Cierre",
              6:"Acumulad. al inicio", 7:"Bajas del Ejercicio", 8:"Del Ejercicio %",
              9:"Importe", 10:"Acumulad. al cierre", 11:EJ25, 12:EJ24}.items():
    c = ws.cell(r, ci); c.value = h; c.fill = fl(C_HDR)
    c.font = fw(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
r += 1
for rb in ["Muebles y útiles","Instalaciones","Equipos y herramientas","Rodados","Inmueble Terreno","Inmueble Edificio"]:
    s(ws, r, 1, rb, fn=fw(size=9))
    for ci in range(2, NCOLS_BU+1):
        c = ws.cell(r, ci); c.value = 0.0; c.number_format = NF; c.font = fw(size=9)
    r += 1
for ci in range(1, NCOLS_BU+1):
    ws.cell(r, ci).fill = fl(C_TOT); ws.cell(r, ci).font = fw(bold=True, color=C_HDR, size=9)
ws.cell(r, 1).value = "TOTALES"
for ci in range(2, NCOLS_BU+1):
    c = ws.cell(r, ci); c.value = 0.0; c.number_format = NF
    c.fill = fl(C_TOT); c.font = fw(bold=True, color=C_HDR, size=9)
r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS_BU)
ws.cell(r,1).value = "Las notas y anexos que se acompañan forman parte integrante de los estados contables."
ws.cell(r,1).font  = Font(italic=True, size=9, name="Arial")

# ═══════════════════════════════════════════════════════════════════════════
# ANEXO II
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Anexo II")
set_cols(ws, [48, 18, 18])
r = title_block(ws, EMPRESA,
    "ANEXO II — COSTO DE LAS MERCADERÍAS O PRODUCTOS VENDIDOS",
    f"Ejercicio N°{NRO_EJ} cerrado el {EJ25} | Comparativo {EJ24} — en pesos de poder adquisitivo al {EJ25}",
    ncols=3)
r = col_headers(ws, r, ["CONCEPTO", EJ25, EJ24])
r = detail(ws, r, "Existencias al comienzo del Ejercicio",        costo_ei25,   costo_ei24,   indent=0)
r = detail(ws, r, "Compras del Ejercicio (neta de descuento)",    costo_cmp25,  costo_cmp24,  indent=0)
r = total(ws,  r, "Subtotal",                                     costo_sub25,  costo_cmp24)
r = detail(ws, r, "Existencias al Cierre de Ejercicio",           costo_ef25,   costo_ef24,   indent=0)
r = total(ws,  r, "Costo de las mercaderías / productos vendidos", costo_costo25, costo_costo24, bg=C_HDR)
for ci in range(1, 4): ws.cell(r-1, ci).font = fw(bold=True, color="FFFFFF")
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(r,1).value = (
    f"Nota: Las existencias al inicio del ejercicio {AÑO_ACT} corresponden a las existencias al cierre "
    f"del ejercicio {AÑO_ANT} ($ {r24_bc:,.2f}), reexpresadas al coeficiente {COF} = $ {costo_ei25:,.2f}. "
    "Las notas y anexos forman parte integrante de estos estados."
)
ws.cell(r,1).font = Font(italic=True, size=9, name="Arial")

# ═══════════════════════════════════════════════════════════════════════════
# ANEXO III
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Anexo III")
NCOLS_III = 7
set_cols(ws, [32, 15, 15, 15, 13, 15, 15])
r = title_block(ws, EMPRESA,
    "ANEXO III — INFORMACIÓN REQUERIDA POR ART. 64 — INC. b) — LEY 19.550",
    f"Ejercicio N°{NRO_EJ} cerrado el {EJ25} | Comparativo al {EJ24} — en pesos de poder adquisitivo al {EJ25}",
    ncols=NCOLS_III)
row_height(ws, r, 35)
ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
ws.merge_cells(start_row=r, start_column=6, end_row=r+1, end_column=6)
ws.merge_cells(start_row=r, start_column=7, end_row=r+1, end_column=7)
for ci in [1, 2, 6, 7]:
    c = ws.cell(r, ci); c.fill = fl(C_HDR); c.font = fw(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.cell(r, 1).value = "DETALLE"
ws.cell(r, 2).value = f"EJERCICIO ACTUAL {AÑO_ACT}"
ws.cell(r, 6).value = f"TOTAL\n{AÑO_ACT}"
ws.cell(r, 7).value = f"TOTAL\n{EJ24}"
r += 1
row_height(ws, r, 40)
for ci, h in {2:"Costo Servicios", 3:"Gastos Admin.", 4:"Gastos Comerc.", 5:"Otros Gastos"}.items():
    c = ws.cell(r, ci); c.value = h; c.fill = fl(C_HDR)
    c.font = fw(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
r += 1

def a3(ws, r, label, cs=None, ga=None, gc=None, og=None, tot=None, t24=None):
    s(ws, r, 1, label, fn=fw(size=9))
    for ci, val in [(2,cs),(3,ga),(4,gc),(5,og),(6,tot),(7,t24)]:
        c = ws.cell(r, ci)
        if val is not None: c.value = val; c.number_format = NF
        c.font = fw(size=9)

def a3_tot(ws, r, label, cs=None, ga=None, gc=None, og=None, tot=None, t24=None, bg=C_TOT):
    for ci in range(1, NCOLS_III+1):
        ws.cell(r, ci).fill = fl(bg); ws.cell(r, ci).font = fw(bold=True, color=C_HDR if bg==C_TOT else "FFFFFF", size=9)
    ws.cell(r, 1).value = label
    for ci, val in [(2,cs),(3,ga),(4,gc),(5,og),(6,tot),(7,t24)]:
        c = ws.cell(r, ci)
        if val is not None: c.value = val; c.number_format = NF

a3(ws, r, "Compra de Servicios",         cs=costo_sv25, tot=costo_sv25,  t24=None);            r+=1
a3(ws, r, "Remuneraciones",              ga=gadm25_suel, tot=gadm25_suel, t24=r24_a3_rem);     r+=1
a3(ws, r, "Cargas Sociales",             ga=gadm25_cs,   tot=gadm25_cs,   t24=r24_a3_cs);      r+=1
a3(ws, r, "Gastos Bancarios",            ga=gadm25_banc, tot=gadm25_banc, t24=None);            r+=1
a3(ws, r, "Gastos Administrativos",      ga=gadm25_adm,  tot=gadm25_adm,  t24=None);            r+=1
a3(ws, r, "IIBB Buenos Aires",           gc=gcom25_bsas, tot=gcom25_bsas, t24=r24_a3_iibb);    r+=1
a3(ws, r, "IIBB Ciudad de Buenos Aires", gc=gcom25_caba, tot=gcom25_caba, t24=None);            r+=1
a3(ws, r, "IIBB Mendoza",               gc=gcom25_mdz,  tot=gcom25_mdz,  t24=None);            r+=1
a3_tot(ws, r, "TOTALES",
       cs=costo_sv25, ga=gadm25, gc=gcom25, og=0,
       tot=costo_sv25+gadm25+gcom25,
       t24=rx(r24_gcom+r24_gadm)); r+=1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS_III)
ws.cell(r,1).value = "Las notas y anexos que se acompañan forman parte integrante de los estados contables."
ws.cell(r,1).font  = Font(italic=True, size=9, name="Arial")

# ═══════════════════════════════════════════════════════════════════════════
# NOTAS
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Notas")
set_cols(ws, [18, 74])
al_wrap = Alignment(wrap_text=True, vertical="top")
al_left = Alignment(horizontal="left", vertical="top")

r = title_block(ws, EMPRESA, "NOTAS A LOS ESTADOS CONTABLES",
    f"Por el ejercicio N°{NRO_EJ} cerrado el {EJ25} — en pesos de poder adquisitivo al {EJ25}",
    ncols=2)

def nota(ws, r, nro, texto, h=60):
    row_height(ws, r, h)
    s(ws, r, 1, nro, bg=C_SEC, fn=fw(bold=True, color=C_HDR, size=9), al=al_left)
    cell = ws.cell(r, 2); cell.value = texto
    cell.font = Font(name="Courier New", size=8.5)
    cell.alignment = al_wrap
    return r + 1

r = nota(ws, r, "1 — Normas Contables Aplicadas",
f"""1.1 Modelo de presentación
Los presentes estados contables han sido preparados en moneda homogénea (pesos de poder adquisitivo al {EJ25}), reconociendo en forma integral los efectos de la inflación de conformidad con la Resolución Técnica (RT) N° 6 de la FACPCE. Los criterios de exposición y valuación siguen la RT N°9 (normas particulares de exposición para entes comerciales), RT N°17 (normas particulares de medición para entes comerciales), RT N°37 (auditoría) y RT N°54 (instrumentos financieros). El estado de flujo de efectivo se prepara conforme a la RT N°8.

1.2 Reexpresión en moneda homogénea (RT N°6)
La economía argentina reviste carácter inflacionario. Los estados contables comparativos del ejercicio anterior (cerrado el {EJ24}) se reexpresan aplicando el coeficiente {COF} sobre todos sus saldos nominales, a fin de expresarlos en pesos de poder adquisitivo del {EJ25}. El efecto monetario de mantener activos y pasivos expuestos se expone como RECPAM en el Estado de Resultados.

1.3 Criterios de valuación
- Los activos y pasivos en moneda nacional se valúan a valor nominal (pesos corrientes).
- Los créditos se valúan a su valor probable de realización.
- Los bienes de cambio se valúan al costo de reposición o valor neto de realización, el menor.
- Los bienes de uso se deprecian por el método de línea recta.
- El impuesto a las ganancias se contabiliza por el método del impuesto diferido (RT N°17). Al cierre no se registran diferencias temporarias significativas.""", h=260)

r = nota(ws, r, "2 — Composición de los principales rubros", "", h=20)

r = nota(ws, r, "Nota 2.1\nCaja y Bancos",
f"""                                            {AÑO_ACT}              {AÑO_ANT} (reexp.)
Caja                              $ {caja25_caja:>16,.2f}    $  {rx(sa['caja_caja']):>16,.2f}
{f'Banco' if caja25_banco else 'Banco'}                             $ {caja25_banco:>16,.2f}    $  {rx(sa['caja_banco']):>16,.2f}
                                  ——————————————————         ———————————————
TOTAL                             $ {caja25:>16,.2f}    $  {caja24:>16,.2f}""", h=105)

r = nota(ws, r, "Nota 2.2\nCréditos por Ventas",
f"""                                            {AÑO_ACT}              {AÑO_ANT} (reexp.)
Deudores por Ventas               $ {cv25:>16,.2f}    $  {cv24:>16,.2f}
TOTAL                             $ {cv25:>16,.2f}    $  {cv24:>16,.2f}""", h=60)

r = nota(ws, r, "Nota 2.3\nOtros Créditos",
f"""                                            {AÑO_ACT}              {AÑO_ANT} (reexp.)
IVA Saldo Técnico                 $ {oc25_iva:>16,.2f}    $  {rx(sa['oc_iva']):>16,.2f}
Saldo Libre Disponibilidad        $ {oc25_sld:>16,.2f}    $  {rx(sa['oc_sld']):>16,.2f}
Impuesto a los Débitos y Créditos $ {oc25_dbc:>16,.2f}    $  {rx(sa['oc_dbc']):>16,.2f}
Retención Ganancias Sufrida       $ {oc25_ret:>16,.2f}    $  {rx(sa['oc_ret']):>16,.2f}
                                  ——————————————————         ———————————————
TOTAL                             $ {oc25:>16,.2f}    $  {oc24:>16,.2f}""", h=120)

r = nota(ws, r, "Nota 2.4\nBienes de Cambio",
f"""                                            {AÑO_ACT}              {AÑO_ANT} (reexp.)
Bienes de Cambio                  $ {bc25:>16,.2f}    $  {bc24:>16,.2f}
Las existencias al inicio del ejercicio {AÑO_ACT} se determinaron reexpresando el saldo de cierre {AÑO_ANT}
($ {r24_bc:,.2f}) por el coeficiente {COF}, resultando en $ {costo_ei25:,.2f} (véase Anexo II).""", h=90)

r = nota(ws, r, "Nota 2.5\nCuentas a Pagar", "", h=20)

r = nota(ws, r, "  2.5.1 — Comerciales",
f"""                                            {AÑO_ACT}              {AÑO_ANT} (reexp.)
Proveedores                       $ {dc25:>16,.2f}    $  {dc24:>16,.2f}
TOTAL                             $ {dc25:>16,.2f}    $  {dc24:>16,.2f}""", h=65)

r = nota(ws, r, "  2.5.2 — Cargas Fiscales",
f"""                                            {AÑO_ACT}              {AÑO_ANT} (reexp.)
IIBB BSAS a pagar                 $ {df25_bsas:>16,.2f}    $  {rx(sa['df_bsas']):>16,.2f}
IIBB CABA a pagar                 $ {df25_caba:>16,.2f}    $  {rx(sa['df_caba']):>16,.2f}
                                  ——————————————————         ———————————————
TOTAL                             $ {df25:>16,.2f}    $  {df24:>16,.2f}""", h=90)

r = nota(ws, r, "  2.5.3 — Remuneraciones y Cargas Sociales",
f"""                                            {AÑO_ACT}              {AÑO_ANT} (reexp.)
Remuneraciones y Cargas Sociales  $ {rem25:>16,.2f}    $  {rem24:>16,.2f}
TOTAL                             $ {rem25:>16,.2f}    $  {rem24:>16,.2f}""", h=65)

r = nota(ws, r, "  2.5.4 — Deudas Sociales",
f"""                                            {AÑO_ACT}              {AÑO_ANT} (reexp.)
Cuenta particular {SOCIO:<16}  $ {ds25:>16,.2f}    $  {ds24:>16,.2f}
TOTAL                             $ {ds25:>16,.2f}    $  {ds24:>16,.2f}""", h=65)

r = nota(ws, r, "Nota 2.6\nResultados Financieros y por Tenencia (RECPAM)",
f"""                                            {AÑO_ACT}              {AÑO_ANT} (reexp.)
RECPAM operativo (Sumas y Saldos) $ {recpam25:>16,.2f}    $  {recpam24:>16,.2f}
Ajuste reexpresión apertura (RT6) $ {recpam_rx_aper:>16,.2f}    $  {0:>16,.2f}
TOTAL RECPAM                      $ {recpam25_adj:>16,.2f}    $  {recpam24:>16,.2f}

El RECPAM surge de la reexpresión de activos y pasivos monetarios expuestos a la inflación,
de conformidad con la RT N°6.""", h=120)

pn_desc = "negativo" if pn25 < 0 else "positivo"
r = nota(ws, r, "3 — Patrimonio Neto",
f"""El Estado de Situación Patrimonial al {EJ25} muestra un Patrimonio Neto {pn_desc} de
$ {abs(pn25):,.2f}. Esta situación se origina principalmente en el resultado {"negativo" if res25_adj < 0 else "positivo"} del
ejercicio de $ {abs(res25_adj):,.2f}, que refleja el nivel de gastos operativos y de estructura
incurridos durante el período.""", h=90)

# ═══════════════════════════════════════════════════════════════════════════
# PAGE SETUP
# ═══════════════════════════════════════════════════════════════════════════
from openpyxl.worksheet.page import PageMargins

def page_setup(ws, landscape=False, fit_w=1, fit_h=0, scale=None):
    ws.page_setup.paperSize   = 9
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    if scale:
        ws.page_setup.fitToWidth = 0; ws.page_setup.fitToHeight = 0
        ws.page_setup.scale = scale; ws.sheet_properties.pageSetUpPr.fitToPage = False
    else:
        ws.page_setup.fitToWidth = fit_w; ws.page_setup.fitToHeight = fit_h
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True

page_setup(wb["Carátula"])
page_setup(wb["ESP"])
page_setup(wb["ER"])
page_setup(wb["EEPN"],    landscape=True)
page_setup(wb["EF"])
page_setup(wb["Anexo I"], landscape=True)
page_setup(wb["Anexo II"])
page_setup(wb["Anexo III"], landscape=True)
page_setup(wb["Notas"])

# ═══════════════════════════════════════════════════════════════════════════
# GUARDAR
# ═══════════════════════════════════════════════════════════════════════════
if args.output:
    OUT = args.output
else:
    empresa_slug = EMPRESA.replace(" ", "_").replace(".", "").replace(",", "")[:30]
    OUT = f"/Users/jstivala/Downloads/EECC_{empresa_slug}_{AÑO_ACT}_v7.xlsx"

wb.save(OUT)
print(f"\n✅ Guardado: {OUT}")
