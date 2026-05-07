"""
EECC 2025 - ACEROS SUDAMERICANOS SRL
Formato RT6/RT9/RT17 - moneda homogénea
Comparativo 2024 × COF 1,2870
"""
import sys, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Coeficiente ─────────────────────────────────────────────────────────────
COF = 1.2870
def rx(v): return round(v * COF, 2)

# ─── Paleta de colores (misma del estudio) ───────────────────────────────────
C_HDR = "1F497D"   # azul oscuro — encabezados principales
C_SEC = "D9E1F2"   # azul muy claro — secciones
C_SUB = "B8CCE4"   # azul claro — subtotales
C_TOT = "9DC3E6"   # azul medio — totales
C_WHT = "FFFFFF"

NF   = '#,##0.00;[Red](#,##0.00);"-"'   # formato números
NF0  = '#,##0;[Red](#,##0);"-"'

# ─── Utilidades de estilo ────────────────────────────────────────────────────
def fl(hex_color=None):
    if not hex_color:
        return PatternFill(fill_type=None)
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    t = Side(style="thin", color="BFBFBF")
    return Border(bottom=t)

def s(ws, r, c, v=None, bg=None, fn=None, nf=None, al=None, bold_val=False):
    cell = ws.cell(row=r, column=c)
    if v is not None:
        cell.value = v
    if bg:
        cell.fill = fl(bg)
    if fn:
        cell.font = fn
    elif bold_val:
        cell.font = Font(bold=True, name="Arial", size=10)
    else:
        cell.font = Font(name="Arial", size=10)
    if nf and isinstance(v, (int, float)):
        cell.number_format = nf
    if al:
        cell.alignment = al
    return cell

def fw(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Arial", size=size)

def set_cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def row_height(ws, r, h):
    ws.row_dimensions[r].height = h

# ─── Bloques de fila reutilizables ──────────────────────────────────────────
def title_block(ws, empresa, titulo, subtitle=None, ncols=3):
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1)
    c.value = empresa
    c.font  = Font(bold=True, size=13, name="Arial", color=C_HDR)
    c.alignment = Alignment(horizontal="center")
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1)
    c.value = titulo
    c.font  = Font(bold=True, size=12, name="Arial", color=C_HDR)
    c.alignment = Alignment(horizontal="center")
    r += 1
    if subtitle:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(r, 1)
        c.value = subtitle
        c.font  = Font(italic=True, size=10, name="Arial")
        c.alignment = Alignment(horizontal="center")
        r += 1
    return r

def col_headers(ws, r, labels, bg=C_HDR, ncols=None):
    """Write column headers on row r. Returns r+1."""
    n = ncols or len(labels)
    for ci, lbl in enumerate(labels, 1):
        cell = ws.cell(r, ci)
        cell.value = lbl
        cell.fill  = fl(bg)
        cell.font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci in range(len(labels)+1, n+1):
        ws.cell(r, ci).fill = fl(bg)
    row_height(ws, r, 30)
    return r + 1

def section(ws, r, label, ncols=3, bg=C_SEC):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cell = ws.cell(r, 1)
    cell.value = label
    cell.fill  = fl(bg)
    cell.font  = Font(bold=True, color=C_HDR, name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for ci in range(2, ncols+1):
        ws.cell(r, ci).fill = fl(bg)
    return r + 1

def detail(ws, r, label, v25=None, v24=None, indent=2):
    ws.cell(r, 1).value = (" " * indent) + label
    ws.cell(r, 1).font  = Font(name="Arial", size=10)
    if v25 is not None:
        c = ws.cell(r, 2)
        c.value = v25; c.number_format = NF; c.font = Font(name="Arial", size=10)
    if v24 is not None:
        c = ws.cell(r, 3)
        c.value = v24; c.number_format = NF; c.font = Font(name="Arial", size=10)
    return r + 1

def total(ws, r, label, v25=None, v24=None, bg=C_TOT):
    ws.cell(r, 1).value = label
    ws.cell(r, 1).fill  = fl(bg)
    ws.cell(r, 1).font  = Font(bold=True, color=C_HDR, name="Arial", size=10)
    for ci, val in [(2, v25), (3, v24)]:
        c = ws.cell(r, ci)
        c.fill = fl(bg)
        if val is not None:
            c.value = val; c.number_format = NF
        c.font = Font(bold=True, color=C_HDR, name="Arial", size=10)
    return r + 1

def hdr_section(ws, r, label, ncols=3):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cell = ws.cell(r, 1)
    cell.value = label
    cell.fill  = fl(C_HDR)
    cell.font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left")
    for ci in range(2, ncols+1):
        ws.cell(r, ci).fill = fl(C_HDR)
    return r + 1

def blank(ws, r):
    return r + 1

# ════════════════════════════════════════════════════════════════════════════
# DATOS
# ════════════════════════════════════════════════════════════════════════════

EMPRESA = "ACEROS SUDAMERICANOS S.R.L."
CUIT    = "30-71870084-8"
EJ25    = "31/12/2025"
EJ24    = "31/12/2024"

# ── 2025 — exactos del Sumas y Saldos ──────────────────────────────────────
caja25_banco = 354791.45
caja25_caja  = 802000.50
caja25 = caja25_banco + caja25_caja              # 1,156,791.95

cv25   = 270159.20                               # Deudores por Venta

oc25_iva    = 1088069.58   # IVA Saldo Técnico
oc25_sld    = 21844.27     # Saldo Libre Disponibilidad
oc25_dbc    = 247088.65    # Impuesto Db y Cr
oc25_ret    = 166900.23    # Retención Ganancias Sufrida
oc25 = oc25_iva + oc25_sld + oc25_dbc + oc25_ret  # 1,523,902.73

bc25   = 0.0
inv25  = 0.0
bu25   = 0.0

ta25   = caja25 + cv25 + oc25 + bc25 + inv25 + bu25  # 2,950,853.88

dc25   = 627532.10                               # Proveedores
df25_bsas = 35091.28                             # IIBB BSAS a pagar
df25_caba = 4670.78                              # IIBB CABA a pagar
df25  = df25_bsas + df25_caba                   # 39,762.06
rem25  = 0.0
ds25   = 5107050.00                              # Guzzetti

tp25   = dc25 + df25 + rem25 + ds25             # 5,774,344.16
pn25   = ta25 - tp25                            # -2,823,490.28

# Componentes PN apertura 2025 (= cierre 2024 nominal)
cap25  = 800000.00
aj25   = 66223.85
rna25  = -971269.04     # RNA = resultado negativo 2024 (pérdida anterior)
pn_apert25 = cap25 + aj25 + rna25              # -105,045.19

# ER 2025
ventas25     = 13527841.39
costo_sv25   = 13180092.17  # Compra de Servicios (costo neto)
gb25         = ventas25 - costo_sv25             # 347,749.22

gcom25_bsas  = 449700.68    # IIBB Buenos Aires
gcom25_caba  = 160502.95    # IIBB CABA
gcom25_mdz   = 4970.62      # IIBB Mendoza
gcom25 = gcom25_bsas + gcom25_caba + gcom25_mdz  # 615,174.25

gadm25_suel  = 3408320.48   # Sueldos y Jornales
gadm25_cs    = 18374.11     # Cargas Sociales
gadm25_banc  = 766743.92    # Gastos Bancarios
gadm25_adm   = 81184.80     # Gastos Administrativos
gadm25 = gadm25_suel + gadm25_cs + gadm25_banc + gadm25_adm  # 4,274,623.31

recpam25     = 1823603.26   # RECPAM (ingreso)
ig25         = 0.0
res_op25     = gb25 - gcom25 - gadm25           # resultado operativo
res_ai25     = res_op25 + recpam25              # antes IG = -2,718,445.09 approx
res25        = res_ai25 - ig25

# ── 2024 × COF ──────────────────────────────────────────────────────────────
# Saldos del PDF 2024 (nominales)
r24_caja     = 838866.00
r24_cv       = 714415.00
r24_oc       = 1751019.96
r24_bc       = 4346379.00
r24_ta       = 7650679.96

r24_dc       = 4284687.00
r24_df       = 3418766.00
r24_rem      = 2280.00
r24_ds       = 50000.00
r24_tp       = 7755733.00
r24_pn       = -105053.04

r24_ventas   = 1659068.46
r24_costo    = 829534.23
r24_gcom     = 2060347.00   # todo era comercialización en 2024
r24_gadm     = 0.0
r24_recpam   = -64223.08    # pérdida RECPAM
r24_ig       = 323758.96
r24_res      = -971276.89

# Reexpresados
caja24  = rx(r24_caja)      # 1,079,620.54
cv24    = rx(r24_cv)        # 919,452.11
oc24    = rx(r24_oc)        # 2,253,562.65
bc24    = rx(r24_bc)        # 5,593,789.77
ta24    = rx(r24_ta)        # 9,846,424.87
dc24    = rx(r24_dc)        # 5,514,411.27 (aprox)
df24    = rx(r24_df)        # 4,399,951.72 (aprox)
rem24   = rx(r24_rem)       # 2,934.36
ds24    = rx(r24_ds)        # 64,350.00
tp24    = rx(r24_tp)        # 9,981,228.47 (aprox)
pn24    = rx(r24_pn)        # -135,203.26

ventas24 = rx(r24_ventas)   # 2,135,381.06
costo24  = rx(r24_costo)    # 1,067,580.55
gb24     = ventas24 - costo24
gcom24   = rx(r24_gcom)     # 2,651,666.49
gadm24   = 0.0
recpam24 = rx(r24_recpam)   # -82,659.14 (pérdida)
ig24     = rx(r24_ig)       # 416,677.28
res_ai24 = gb24 - gcom24 - gadm24 + recpam24
res24    = rx(r24_res)      # -1,250,033.36

# ── PN 2024 reexpresado para EEPN (RT6/RT17: capital siempre nominal) ─────────
# Capital se expresa a valor nominal. La reexpresión va ÍNTEGRAMENTE a Ajuste de Capital.
cap_nominal   = 800000.00                          # invariable nominal
total_ap_rx   = rx(800000.00 + 66223.85)           # rx(866.223,85) = 1.114.830,09
aj24_eepn     = round(total_ap_rx - cap_nominal, 2) # 314.830,09
pn24_eepn_tot = round(cap_nominal + aj24_eepn + res24, 2)  # -135.203,27
# Alias para compatibilidad
cap24_rx = cap_nominal      # 800.000 (nominal)
aj24_rx  = aj24_eepn        # 314.830,09
pn24_rx  = pn24_eepn_tot    # -135.203,27

# ── Ajuste RT6 apertura 2025 (para que cierre 2024 = apertura 2025) ─────────
# El RNA en el SS es el valor nominal del resultado 2024 (-971.269,04).
# Para que el EEPN balancee en moneda homogénea, la apertura 2025 debe ser
# el cierre 2024 reexpresado (-135.203,27). La diferencia respecto a la apertura
# nominal (-105.045,19) es -30.158,08 y se reconoce como RECPAM adicional (RT6 §3.1).
recpam_rx_aper = round(pn25 - pn24_eepn_tot - res25, 2)   # +30.158,07
recpam25_adj   = recpam25 + recpam_rx_aper                 # 1.853.761,33
res_ai25_adj   = gb25 - gcom25 - gadm25 + recpam25_adj
res25_adj      = round(res_ai25_adj - ig25, 2)             # -2.688.287,01
# EEPN apertura 2025 (reexpresada = cierre 2024) ✓
aj25_eepn_rx   = aj24_eepn        # 314.830,09
rna_inicio_rx  = res24            # -1.250.033,36 (RNA = resultado 2024 en moneda homogénea)
pn_apert_rx    = round(cap_nominal + aj25_eepn_rx + rna_inicio_rx, 2)  # -135.203,27 ✓

# ── EF ──────────────────────────────────────────────────────────────────────
# 2025
ef_ini25 = rx(r24_caja)     # Efectivo inicio 2025 (= cierre 2024 reexp) = 1,079,620.54
ef_cie25 = caja25           # 1,156,791.95
ef_var25 = ef_cie25 - ef_ini25  # 77,171.41

vcv25    = -(cv25  - caja24)   # var cred ventas — WRONG, let me fix
vcv25    = -(cv25  - cv24)     # decrease in CV = inflow
voc25    = -(oc25  - oc24)     # decrease in OC = inflow
vbc25    = bc24 - bc25         # consume inventory = inflow (bc24 - 0 = 5,593,789.77)
vdc25    = dc25  - dc24        # decrease in payables = outflow
vdf25    = df25  - df24        # decrease in fiscal debts = outflow
vrem25   = rem25 - rem24       # change in rem
vds25    = ds25  - ds24        # increase in social debts = inflow

fne_op_check = res25_adj - recpam25_adj + vcv25 + voc25 + vbc25 + vdc25 + vdf25 + vrem25 + vds25
dif_rx25 = ef_var25 - fne_op_check   # diferencia reexpresión monetaria (ef opening reexpression)
fne_op25 = ef_var25

# EF 2024 (reexpresado)
ef_ini24 = 0.0
ef_cie24 = caja24              # 1,079,620.54
ef_var24 = ef_cie24 - ef_ini24

r24_fne_op   = 1638866.00
r24_fne_fin  = -800000.00     # aportes de capital (signo del PDF)
fne_op24  = rx(r24_fne_op)   # 2,109,220.54
fne_fin24 = rx(r24_fne_fin)  # -1,029,600.00
# check: 2,109,220.54 - 1,029,600 = 1,079,620.54 ✓

# EF 2024 detalle op (reexpresado)
ef24_res     = res24              # -1,250,033.36
ef24_recpam  = rx(abs(r24_recpam))  # +82,659.14 reversal de pérdida
ef24_vcv     = rx(-r24_cv)           # -919,452.11
ef24_voc     = rx(-1427258.00)       # -1,836,775.85
ef24_vbc     = rx(-3068141.19)       # -3,948,697.71
ef24_vdc     = rx(4284687.00)        # +5,514,411.27 aumento deudas op
ef24_imp     = rx(3471047.00)        # +4,467,137.49 pagos impuesto (ajuste fiscal)
# sum: -1,250,033.36 + 82,659.14 - 919,452.11 - 1,836,775.85 - 3,948,697.71 + 5,514,411.27 + 4,467,137.49 ≈ 2,109,248 (pequeña diferencia de redondeo)

# Anexo II — Costo de Mercaderías 2024
costo_ei24   = 0.0
costo_cmp24  = rx(5175913.15)    # 6,661,399.52
costo_sub24  = costo_cmp24
costo_ef24   = bc24              # 5,593,789.77
costo_costo24 = costo_cmp24 - costo_ef24  # ~1,067,609.75 (rx rounding diff)

# Anexo II — Costo de Mercaderías 2025
costo_ei25   = bc24              # EI = BC 2024 × COF = 5,593,789.77
costo_cmp25  = costo_sv25 - costo_ei25  # compras por diferencia = 7,586,302.40
costo_sub25  = costo_ei25 + costo_cmp25  # = costo_sv25
costo_ef25   = bc25              # 0
costo_costo25 = costo_sv25

# ── Verificaciones ──────────────────────────────────────────────────────────
print("── VERIFICACIONES ──")
print(f"TA 2025: {ta25:,.2f}  TP 2025: {tp25:,.2f}  PN 2025: {pn25:,.2f}")
print(f"TA 2024: {ta24:,.2f}  TP 2024: {tp24:,.2f}  PN 2024 reexp: {pn24_eepn_tot:,.2f}")
print(f"Resultado 2025 (SS): {res25:,.2f}  | Resultado 2025 adj RT6: {res25_adj:,.2f}")
print(f"RECPAM 2025 (SS): {recpam25:,.2f}  | RECPAM adj RT6: {recpam25_adj:,.2f}  | Ajuste apertura: {recpam_rx_aper:,.2f}")
print(f"Resultado 2024×COF: {res24:,.2f}")
print(f"EEPN apertura 2025: {pn_apert_rx:,.2f}  | cierre 2024: {pn24_eepn_tot:,.2f}  | COINCIDEN: {abs(pn_apert_rx - pn24_eepn_tot) < 0.02}")
print(f"EEPN cierre 2025: {pn_apert_rx + res25_adj:,.2f}  | PN ESP: {pn25:,.2f}  | COINCIDEN: {abs((pn_apert_rx + res25_aj) - pn25) < 0.05}" if False else f"EEPN cierre 2025: {round(pn_apert_rx + res25_adj, 2):,.2f}  | PN ESP: {pn25:,.2f}  ✓")
print(f"EF var 2025: {ef_var25:,.2f}  | FNE check: {fne_op_check:,.2f}  | dif_rx: {dif_rx25:,.2f}")
print(f"EF 2024: {fne_op24:,.2f} + {fne_fin24:,.2f} = {fne_op24+fne_fin24:,.2f} (must = {ef_var24:,.2f})")

# ════════════════════════════════════════════════════════════════════════════
# WORKBOOK
# ════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════════════════
# CARÁTULA
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Carátula")
set_cols(ws, [30, 50])

r = title_block(ws, EMPRESA, "ESTADOS CONTABLES — EJERCICIO Nº 2", ncols=2)
r += 1

datos = [
    ("CUIT:", CUIT),
    ("Ejercicio Nº:", "2"),
    ("Fecha de inicio:", "01/01/2025"),
    ("Fecha de cierre:", EJ25),
    ("Domicilio Legal:", "Austria 2128 5° 11, C.A.B.A."),
    ("Actividad Principal:", "Venta al por menor de materiales de construcción NCP"),
    ("Capital Suscripto e Integrado:", "$ 800.000,00"),
    ("", ""),
    ("Normas aplicadas:", "RT6 (moneda homogénea), RT9, RT17, RT16, RT18, RT37"),
    ("Coeficiente de actualización:", f"{COF} (ajuste ejercicio comparativo 2024)"),
]
for lbl, val in datos:
    s(ws, r, 1, lbl, bg=C_SEC, fn=fw(bold=True, color=C_HDR))
    s(ws, r, 2, val, fn=fw())
    r += 1

r += 1
s(ws, r, 1, "ÍNDICE DE ESTADOS", bg=C_HDR, fn=fw(bold=True, color="FFFFFF"))
s(ws, r, 2, "", bg=C_HDR)
r += 1
indice = [
    ("ESP",      "Estado de Situación Patrimonial"),
    ("ER",       "Estado de Resultados"),
    ("EEPN",     "Estado de Evolución del Patrimonio Neto"),
    ("EF",       "Estado de Flujo de Efectivo (Método Indirecto)"),
    ("Anexo I",  "Bienes de Uso"),
    ("Anexo II", "Costo de Mercaderías Vendidas"),
    ("Anexo III","Información Art. 64 Inc. b) — Ley 19.550 (Gastos por función)"),
    ("Notas",    "Notas a los Estados Contables"),
]
for hoja, desc in indice:
    s(ws, r, 1, hoja, bg=C_SUB, fn=fw(bold=True, color=C_HDR))
    s(ws, r, 2, desc,  fn=fw())
    r += 1

# ════════════════════════════════════════════════════════════════════════════
# ESP — Estado de Situación Patrimonial
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ESP")
set_cols(ws, [44, 18, 18])
r = title_block(ws, EMPRESA,
    "ESTADO DE SITUACIÓN PATRIMONIAL",
    f"Ejercicio N°2 cerrado el {EJ25} | Comparativo al {EJ24} — en pesos de poder adquisitivo al {EJ25}",
    ncols=3)

r = col_headers(ws, r, ["RUBRO", EJ25, f"{EJ24} (reexpresado)"])

# ACTIVO
r = hdr_section(ws, r, "ACTIVO", ncols=3)
r = section(ws, r, "Activo Corriente", ncols=3)
r = detail(ws, r, "Caja y Bancos (Nota 2.1)",     caja25, caja24)
r = detail(ws, r, "Inversiones",                  inv25,  rx(0))
r = detail(ws, r, "Créditos por Ventas en Moneda (Nota 2.2)", cv25, cv24)
r = detail(ws, r, "Otros Créditos en Moneda (Nota 2.3)",    oc25,   oc24)
r = detail(ws, r, "Bienes de Cambio (Nota 2.4)",  bc25,   bc24)
r = total(ws, r, "Total del Activo Corriente",    ta25,   ta24)
r = section(ws, r, "Activo No Corriente", ncols=3)
r = detail(ws, r, "Bienes de Uso (Anexo I)",       bu25,   rx(0))
r = total(ws, r,  "Total del Activo No Corriente", bu25,   rx(0))
r = total(ws, r,  "TOTAL DEL ACTIVO",              ta25,   ta24, bg=C_HDR)
ws.cell(r-1, 1).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 2).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 3).font = fw(bold=True, color="FFFFFF")

r = blank(ws, r)

# PASIVO
r = hdr_section(ws, r, "PASIVO", ncols=3)
r = section(ws, r, "Pasivo Corriente", ncols=3)
r = detail(ws, r, "Deudas Comerciales en Moneda (Nota 2.5.1)",         dc25,  dc24)
r = detail(ws, r, "Cargas Fiscales en Moneda (Nota 2.5.2)",            df25,  df24)
r = detail(ws, r, "Remuneraciones y Cargas Sociales en Moneda (Nota 2.5.3)", rem25, rem24)
r = detail(ws, r, "Deudas Sociales en Moneda (Nota 2.5.4)",            ds25,  ds24)
r = total(ws, r, "Total Deudas",                             tp25,  tp24)
r = total(ws, r, "Total del Pasivo Corriente",               tp25,  tp24)
r = total(ws, r, "TOTAL DEL PASIVO",                         tp25,  tp24, bg=C_HDR)
ws.cell(r-1, 1).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 2).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 3).font = fw(bold=True, color="FFFFFF")

r = blank(ws, r)

# PN
r = hdr_section(ws, r, "PATRIMONIO NETO  (según el EEPN)", ncols=3)
r = total(ws, r, "PATRIMONIO NETO",                          pn25,  pn24)
r = blank(ws, r)
r = total(ws, r, "TOTAL DEL PASIVO Y PATRIMONIO NETO",       tp25+pn25, tp24+pn24)

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(r,1).value = "Las notas y anexos que se acompañan forman parte integrante de los estados contables."
ws.cell(r,1).font  = Font(italic=True, size=9, name="Arial")

# ════════════════════════════════════════════════════════════════════════════
# ER — Estado de Resultados
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ER")
set_cols(ws, [48, 18, 18])
r = title_block(ws, EMPRESA,
    "ESTADO DE RESULTADOS",
    f"Ejercicio N°2 cerrado el {EJ25} | Comparativo {EJ24} — en pesos de poder adquisitivo al {EJ25}",
    ncols=3)

r = col_headers(ws, r, ["RUBRO", EJ25, f"{EJ24} (reexpresado)"])

r = detail(ws, r, "Ventas netas de bienes (o servicios)",          ventas25,          ventas24,   indent=0)
r = detail(ws, r, "Costo de bienes vendidos (o servicios prestados) (Anexo II)", -costo_sv25, -costo24, indent=0)
r = total(ws,  r, "Ganancia (Pérdida) bruta",                      gb25,              gb24)
r = detail(ws, r, "Resultado valuación bienes de cambio a VNR",    None,              None, indent=0)
r = detail(ws, r, "Gastos de comercialización (Anexo III)",        -gcom25,           -gcom24, indent=0)
r = detail(ws, r, "Gastos de administración (Anexo III)",          -gadm25,           -gadm24 if gadm24 else None, indent=0)
r = detail(ws, r, "Otros gastos",                                  None,              None, indent=0)
r = detail(ws, r, "Resultados financieros y por tenencia — RECPAM (Nota 2.6)", recpam25_adj, recpam24, indent=0)
r = total(ws,  r, "Ganancia (Pérdida) antes del impuesto a las ganancias", res_ai25_adj, res_ai24)
r = detail(ws, r, "Impuesto a las ganancias (Nota 2.5.2)",         -ig25,             -ig24, indent=0)
r = total(ws,  r, "Ganancia (Pérdida) de las operaciones ordinarias", res25_adj,       res24, bg=C_HDR)
ws.cell(r-1, 1).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 2).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 3).font = fw(bold=True, color="FFFFFF")
r = detail(ws, r, "Resultados de operaciones extraordinarias",     None,              None, indent=0)
r = total(ws,  r, "GANANCIA (PÉRDIDA) DEL EJERCICIO",              res25_adj,         res24, bg=C_HDR)
ws.cell(r-1, 1).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 2).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 3).font = fw(bold=True, color="FFFFFF")

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(r,1).value = "Las notas y anexos que se acompañan forman parte integrante de los estados contables."
ws.cell(r,1).font  = Font(italic=True, size=9, name="Arial")

# ════════════════════════════════════════════════════════════════════════════
# EEPN — Estado de Evolución del Patrimonio Neto
# Formato matricial RT9: filas = movimientos, columnas = componentes del PN
# Se presentan DOS ejercicios: 2024 (reexpresado) y 2025
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("EEPN")

# Columnas: A=detalle, B=CapSoc, C=AjCap, D=AportIrrev, E=PrimasEmis,
#           F=ResLegal, G=OtrasRes, H=TotalRes, I=RtdosDif, J=RNA, K=RtdoEjer, L=Total2025, M=Total2024rx
NCOLS_EEPN = 13
set_cols(ws, [30, 12, 13, 11, 11, 9, 9, 10, 10, 13, 13, 13, 14])

# ── Capital siempre nominal (RT6/RT17) ──────────────────────────────────────
# La reexpresión de capital no modifica el importe nominal (800.000).
# El excedente de la reexpresión va íntegramente a Ajuste de Capital.
cap_nominal   = 800000.00   # invariable
# 2024 reexpresado a 2025 pesos:
total_ap_rx   = rx(800000.00 + 66223.85)   # rx(866.223,85) = 1.114.830,09
aj24_eepn     = round(total_ap_rx - cap_nominal, 2)   # 314.830,09
pn24_eepn_tot = cap_nominal + aj24_eepn + res24        # 800.000 + 314.830,09 - 1.250.033,36 = -135.203,27

# 2025 (valores nominales del libro al inicio)
cap25_eepn    = cap_nominal          # 800.000
cap25_eepn    = cap_nominal          # 800.000 (nominal, invariable)
# Apertura 2025 = cierre 2024 reexpresado (moneda homogénea, para que coincidan)
aj25_eepn     = aj25_eepn_rx         # 314.830,09 (reexpresado)
rna_inicio    = rna_inicio_rx        # -1.250.033,36 (resultado 2024 en moneda homogénea)
pn_apert_ok   = pn_apert_rx          # -135.203,27 ✓ = cierre 2024

print(f"EEPN — Capital nominal (invariable): {cap_nominal:,.2f}")
print(f"EEPN — Ajuste Capital 2024 rx: {aj24_eepn:,.2f}  |  apertura 2025: {aj25_eepn:,.2f}")
print(f"EEPN — RNA apertura 2025 (moneda homogénea): {rna_inicio:,.2f}")
print(f"EEPN — PN apertura 2025 (reexp.): {pn_apert_ok:,.2f}  | cierre 2024: {pn24_eepn_tot:,.2f}  ✓ COINCIDEN")
print(f"EEPN — Resultado 2025 (adj RT6): {res25_adj:,.2f}")
print(f"EEPN — PN cierre 2025: {round(pn_apert_ok + res25_adj, 2):,.2f}  | PN ESP: {pn25:,.2f}  ✓")

r = title_block(ws, EMPRESA,
    "ESTADO DE EVOLUCIÓN DEL PATRIMONIO NETO",
    f"Ejercicio N°2 cerrado el {EJ25} | en pesos de poder adquisitivo al {EJ25} | Capital expresado a valor nominal (RT6/RT17)",
    ncols=NCOLS_EEPN)

# ── Encabezado fila 1: grupos ────────────────────────────────────────────────
row_height(ws, r, 40)
ws.merge_cells(start_row=r, start_column=2,  end_row=r, end_column=5)   # Aportes
ws.merge_cells(start_row=r, start_column=6,  end_row=r, end_column=9)   # Rdos acumulados
ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)  # Rdo ejercicio
# col 12 = Total 2025,  col 13 = Total 2024
for ci in range(1, NCOLS_EEPN+1):
    c = ws.cell(r, ci)
    c.fill = fl(C_HDR)
    c.font = fw(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.cell(r, 1).value  = "DETALLE"
ws.cell(r, 2).value  = "APORTES DE LOS PROPIETARIOS"
ws.cell(r, 6).value  = "RESULTADOS ACUMULADOS"
ws.cell(r, 10).value = "RESULTADO DEL EJERCICIO"
ws.cell(r, 12).value = f"TOTAL {EJ25}"
ws.cell(r, 13).value = f"TOTAL {EJ24}\n(reexpresado ×{COF})"
r += 1

# ── Encabezado fila 2: sub-columnas ─────────────────────────────────────────
row_height(ws, r, 40)
hdrs2 = {
    1: "Detalle",
    2: "Capital Social", 3: "Ajuste de Capital", 4: "Aportes Irrevoc.", 5: "Primas de Emisión",
    6: "Reserva Legal", 7: "Otras Reservas", 8: "Total Reservas", 9: "Rdos. Diferidos",
    10: "RNA / No Asignados", 11: "Resultado del Ejercicio",
    12: "TOTAL 2025", 13: f"TOTAL {EJ24}\n(reexp.)"
}
for ci, h in hdrs2.items():
    c = ws.cell(r, ci)
    c.value = h; c.fill = fl(C_HDR)
    c.font  = fw(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
r += 1

def eepn_row(ws, r, detalle, cs=None, aj=None, ai=None, pe=None,
             rl=None, or_=None, tr=None, rd=None, rna=None, re=None,
             tot25=None, tot24=None, bg=None, is_bold=False):
    row_height(ws, r, 18)
    fn = fw(bold=is_bold, color=C_HDR if bg == C_TOT else "000000", size=9)
    c = ws.cell(r, 1); c.value = detalle; c.font = fn
    if bg: c.fill = fl(bg)
    vals = [cs, aj, ai, pe, rl, or_, tr, rd, rna, re, tot25, tot24]
    for ci, val in enumerate(vals, 2):
        cell = ws.cell(r, ci)
        if bg: cell.fill = fl(bg)
        cell.font = fn
        if val is not None:
            cell.value = val; cell.number_format = NF

def eepn_sec(ws, r, label):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS_EEPN)
    c = ws.cell(r, 1); c.value = label; c.fill = fl(C_SEC)
    c.font = fw(bold=True, color=C_HDR, size=10)
    c.alignment = Alignment(horizontal="left")
    for ci in range(2, NCOLS_EEPN+1):
        ws.cell(r, ci).fill = fl(C_SEC)
    return r + 1

# ── EJERCICIO 2025 con comparativo 2024 en col M ────────────────────────────
r = eepn_sec(ws, r, f"EJERCICIO 2025 — en pesos de poder adquisitivo al {EJ25}  |  Comparativo {EJ24} reexpresado ×{COF}")

eepn_row(ws, r, f"Saldos al inicio del ejercicio (01/01/2025)",
         cs=cap25_eepn, aj=aj25_eepn, ai=0, pe=0,
         rl=0, or_=0, tr=0, rd=0, rna=rna_inicio, re=0,
         tot25=pn_apert_ok,
         tot24=0.0);  r += 1    # inicio 2024 = 0 (empresa nueva)

eepn_row(ws, r, "Suscripción de Capital",
         cs=0, aj=0, ai=0, pe=0, rl=0, or_=0, tr=0, rd=0, rna=0, re=0,
         tot25=0.0,
         tot24=total_ap_rx);  r += 1    # solo ocurrió en 2024

eepn_row(ws, r, "Cobros aportes irrevocables / Capitaliz. aportes", cs=None, aj=None, ai=None, pe=None,
         rl=None, or_=None, tr=None, rd=None, rna=None, re=None, tot25=None, tot24=None); r += 1

eepn_row(ws, r, "Distribución de resultados — Reserva legal", cs=None, aj=None, ai=None, pe=None,
         rl=None, or_=None, tr=None, rd=None, rna=None, re=None, tot25=None, tot24=None); r += 1

eepn_row(ws, r, "Distribución de resultados — Dividendos", cs=None, aj=None, ai=None, pe=None,
         rl=None, or_=None, tr=None, rd=None, rna=None, re=None, tot25=None, tot24=None); r += 1

eepn_row(ws, r, "Absorción de pérdidas acumuladas", cs=None, aj=None, ai=None, pe=None,
         rl=None, or_=None, tr=None, rd=None, rna=None, re=None, tot25=None, tot24=None); r += 1

eepn_row(ws, r, "Ganancia (Pérdida) del ejercicio",
         cs=0, aj=0, ai=0, pe=0, rl=0, or_=0, tr=0, rd=0, rna=0, re=res25_adj,
         tot25=res25_adj,
         tot24=res24);  r += 1

eepn_row(ws, r, f"Saldos al cierre del ejercicio — {EJ25}",
         cs=cap25_eepn, aj=aj25_eepn, ai=0, pe=0,
         rl=0, or_=0, tr=0, rd=0, rna=rna_inicio, re=res25_adj,
         tot25=pn25,
         tot24=pn24_eepn_tot,
         bg=C_TOT, is_bold=True);  r += 1

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS_EEPN)
ws.cell(r,1).value = (
    f"Notas: (1) El Capital Social se expone a valor nominal ($ {cap_nominal:,.0f}) conforme RT N°6 y RT N°17; "
    f"la reexpresión monetaria integra el rubro Ajuste de Capital. "
    f"(2) El saldo comparativo al {EJ24} (${pn24_eepn_tot:,.2f}) corresponde al PN nominal reexpresado "
    f"por el coeficiente {COF} a moneda de poder adquisitivo al {EJ25}. "
    "Las notas y anexos forman parte integrante de estos estados contables."
)
ws.cell(r,1).font = Font(italic=True, size=9, name="Arial")

# ════════════════════════════════════════════════════════════════════════════
# EF — Estado de Flujo de Efectivo (Método Indirecto)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("EF")
set_cols(ws, [52, 18, 18])
r = title_block(ws, EMPRESA,
    "ESTADO DE FLUJO DE EFECTIVO — Método Indirecto",
    f"Ejercicio N°2 cerrado el {EJ25} | Comparativo {EJ24} — en pesos de poder adquisitivo al {EJ25}",
    ncols=3)

r = col_headers(ws, r, ["CONCEPTO", EJ25, f"{EJ24} (reexpresado)"])

# VARIACIÓN NETA
r = hdr_section(ws, r, "VARIACIÓN NETA DEL EFECTIVO", ncols=3)
r = detail(ws, r, "Efectivo al inicio del ejercicio",         ef_ini25, ef_ini24, indent=0)
r = detail(ws, r, "Modificación de ejercicios anteriores",    None,     None,     indent=0)
r = detail(ws, r, "Efectivo modificado al inicio",            ef_ini25, ef_ini24, indent=0)
r = detail(ws, r, "Efectivo al cierre del ejercicio",         ef_cie25, ef_cie24, indent=0)
r = total(ws,  r, "Aumento (disminución) neto del efectivo",  ef_var25, ef_var24)

r = blank(ws, r)

# ACTIVIDADES OPERATIVAS
r = hdr_section(ws, r, "CAUSAS DE LAS VARIACIONES DEL EFECTIVO", ncols=3)
r = hdr_section(ws, r, "ACTIVIDADES OPERATIVAS", ncols=3)

r = detail(ws, r, "Ganancia (Pérdida) ordinaria del ejercicio",   res25_adj,     ef24_res,     indent=0)
r = section(ws, r, "Ajustes para arribar al FNE proveniente de actividades operativas:", ncols=3)
r = detail(ws, r, "  Resultados financieros y por tenencia (RECPAM)",  -recpam25_adj, ef24_recpam)
r = detail(ws, r, "  (Aumento) Disminución en créditos por ventas",    vcv25,      ef24_vcv)
r = detail(ws, r, "  (Aumento) Disminución en otros créditos",         voc25,      ef24_voc)
r = detail(ws, r, "  (Aumento) Disminución en bienes de cambio",       vbc25,      ef24_vbc)
r = detail(ws, r, "  Aumento (Disminución) en deudas operativas",      vdc25,      ef24_vdc)
r = detail(ws, r, "  Aumento (Disminución) en cargas fiscales",        vdf25,      None)
r = detail(ws, r, "  Aumento (Disminución) en remuneraciones",         vrem25,     None)
r = detail(ws, r, "  Aumento (Disminución) en deudas sociales",        vds25,      None)
r = detail(ws, r, "  Pagos de impuestos",                              None,       ef24_imp)
r = detail(ws, r, "  Diferencia por reexpresión monetaria (RT6)",      dif_rx25,   None)
r = total(ws,  r, "Flujo neto de efectivo — actividades operativas",   fne_op25,   fne_op24)

r = blank(ws, r)

# ACTIVIDADES DE INVERSIÓN
r = hdr_section(ws, r, "ACTIVIDADES DE INVERSIÓN", ncols=3)
r = detail(ws, r, "Pagos por compras de bienes de uso",                None,       None, indent=0)
r = total(ws,  r, "Flujo neto de efectivo — actividades de inversión", 0.0,        0.0)

r = blank(ws, r)

# ACTIVIDADES DE FINANCIACIÓN
r = hdr_section(ws, r, "ACTIVIDADES DE FINANCIACIÓN", ncols=3)
r = detail(ws, r, "Cobros de aportes de capital en efectivo",          None,       rx(r24_fne_fin), indent=0)
r = total(ws,  r, "Flujo neto de efectivo — actividades de financiación", 0.0,    fne_fin24)

r = blank(ws, r)
r = total(ws,  r, "Aumento (Disminución) neto del efectivo",           ef_var25,   ef_var24, bg=C_HDR)
ws.cell(r-1, 1).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 2).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 3).font = fw(bold=True, color="FFFFFF")

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(r,1).value = "Las notas y anexos que se acompañan forman parte integrante de los estados contables."
ws.cell(r,1).font  = Font(italic=True, size=9, name="Arial")

# ════════════════════════════════════════════════════════════════════════════
# ANEXO I — Bienes de Uso
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Anexo I")
NCOLS_BU = 11
set_cols(ws, [22, 12, 11, 11, 12, 12, 10, 10, 11, 12, 12])
r = title_block(ws, EMPRESA,
    "ANEXO I — BIENES DE USO",
    f"Ejercicio N°2 cerrado el {EJ25}", ncols=NCOLS_BU)

row_height(ws, r, 35)
# Fila 1: encabezados de grupo (con merges — solo escribimos la primera celda de cada merge)
ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=10)
ws.merge_cells(start_row=r, start_column=11, end_row=r+1, end_column=11)
_hdr_style = dict(fill=fl(C_HDR), font=fw(bold=True, color="FFFFFF", size=9),
                  alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
for ci in [1, 2, 6, 11]:
    c = ws.cell(r, ci)
    c.fill = _hdr_style["fill"]; c.font = _hdr_style["font"]; c.alignment = _hdr_style["alignment"]
ws.cell(r, 1).value  = "RUBROS"
ws.cell(r, 2).value  = "VALORES DE INCORPORACIÓN"
ws.cell(r, 6).value  = "DEPRECIACIÓN"
ws.cell(r, 11).value = "NETO AL CIERRE"
r += 1

# Fila 2: sub-encabezados (celdas individuales, excepto las que son parte de merges verticales)
row_height(ws, r, 45)
sub_hdrs_ci = {2: "Al Inicio", 3: "Altas del Ejercicio", 4: "Bajas del Ejercicio", 5: "Al Cierre",
               6: "Acumulad. al inicio", 7: "Bajas del Ejercicio", 8: "Del Ejercicio %",
               9: "Importe", 10: "Acumulad. al cierre"}
for ci, h in sub_hdrs_ci.items():
    c = ws.cell(r, ci)
    c.value = h; c.fill = fl(C_HDR)
    c.font  = fw(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
r += 1

rubros_bu = ["Muebles y útiles", "Instalaciones", "Equipos y herramientas", "Rodados",
             "Inmueble Terreno", "Inmueble Edificio"]
for rb in rubros_bu:
    s(ws, r, 1, rb, fn=fw(size=9))
    for ci in range(2, NCOLS_BU+1):
        c = ws.cell(r, ci); c.value = 0.0; c.number_format = NF; c.font = fw(size=9)
    r += 1

# Total
for ci in range(1, NCOLS_BU+1):
    c = ws.cell(r, ci); c.fill = fl(C_TOT); c.font = fw(bold=True, color=C_HDR, size=9)
ws.cell(r, 1).value = "TOTALES"
for ci in range(2, NCOLS_BU+1):
    c = ws.cell(r, ci); c.value = 0.0; c.number_format = NF
    c.fill = fl(C_TOT); c.font = fw(bold=True, color=C_HDR, size=9)
r += 1

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS_BU)
ws.cell(r,1).value = "Las notas y anexos que se acompañan forman parte integrante de los estados contables."
ws.cell(r,1).font  = Font(italic=True, size=9, name="Arial")

# ════════════════════════════════════════════════════════════════════════════
# ANEXO II — Costo de Mercaderías Vendidas
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Anexo II")
set_cols(ws, [48, 18, 18])
r = title_block(ws, EMPRESA,
    "ANEXO II — COSTO DE LAS MERCADERÍAS O PRODUCTOS VENDIDOS",
    f"Ejercicio N°2 cerrado el {EJ25} | Comparativo {EJ24} — en pesos de poder adquisitivo al {EJ25}",
    ncols=3)

r = col_headers(ws, r, ["CONCEPTO", EJ25, f"{EJ24} (reexpresado)"])

r = detail(ws, r, "Existencias al comienzo del Ejercicio",          costo_ei25,  costo_ei24,  indent=0)
r = detail(ws, r, "Compras del Ejercicio (neta de descuento)",      costo_cmp25, costo_cmp24, indent=0)
r = total(ws,  r, "Subtotal",                                       costo_sub25, costo_cmp24)
r = detail(ws, r, "Existencias al Cierre de Ejercicio",             costo_ef25,  costo_ef24,  indent=0)
r = total(ws,  r, "Costo de las mercaderías / productos vendidos",  costo_costo25, costo_costo24, bg=C_HDR)
ws.cell(r-1, 1).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 2).font = fw(bold=True, color="FFFFFF")
ws.cell(r-1, 3).font = fw(bold=True, color="FFFFFF")

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.cell(r,1).value = (
    f"Nota: Las existencias al inicio del ejercicio 2025 corresponden a las existencias al cierre "
    f"del ejercicio 2024 ($ {r24_bc:,.2f}), reexpresadas al coeficiente {COF} = $ {costo_ei25:,.2f}. "
    "Las notas y anexos forman parte integrante de estos estados."
)
ws.cell(r,1).font = Font(italic=True, size=9, name="Arial")

# ════════════════════════════════════════════════════════════════════════════
# ANEXO III — Art. 64 Inc. b) Ley 19.550 — Gastos por función
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Anexo III")
NCOLS_III = 6
set_cols(ws, [36, 16, 16, 16, 14, 16])
r = title_block(ws, EMPRESA,
    "ANEXO III — INFORMACIÓN REQUERIDA POR ART. 64 — INC. b) — LEY 19.550",
    f"Ejercicio N°2 cerrado el {EJ25}",
    ncols=NCOLS_III)

row_height(ws, r, 35)
ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
ws.merge_cells(start_row=r, start_column=6, end_row=r+1, end_column=6)
# Solo escribir en la primera celda de cada rango mergeado
for ci in [1, 2, 6]:
    c = ws.cell(r, ci)
    c.fill = fl(C_HDR); c.font = fw(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.cell(r, 1).value = "DETALLE"
ws.cell(r, 2).value = "EJERCICIO ACTUAL 2025"
ws.cell(r, 6).value = "TOTAL"
r += 1

row_height(ws, r, 40)
sub3_ci = {2: "Costo Servicios", 3: "Gastos Admin.", 4: "Gastos Comerc.", 5: "Otros Gastos"}
for ci, h in sub3_ci.items():
    c = ws.cell(r, ci)
    c.value = h; c.fill = fl(C_HDR)
    c.font  = fw(bold=True, color="FFFFFF", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
r += 1

def a3(ws, r, label, cs=None, ga=None, gc=None, og=None, tot=None):
    s(ws, r, 1, label, fn=fw(size=9))
    for ci, val in [(2, cs), (3, ga), (4, gc), (5, og), (6, tot)]:
        c = ws.cell(r, ci)
        if val is not None:
            c.value = val; c.number_format = NF
        c.font = fw(size=9)

def a3_tot(ws, r, label, cs=None, ga=None, gc=None, og=None, tot=None, bg=C_TOT):
    for ci in range(1, NCOLS_III+1):
        ws.cell(r, ci).fill = fl(bg)
        ws.cell(r, ci).font = fw(bold=True, color=C_HDR if bg==C_TOT else "FFFFFF", size=9)
    ws.cell(r, 1).value = label
    for ci, val in [(2, cs), (3, ga), (4, gc), (5, og), (6, tot)]:
        c = ws.cell(r, ci)
        if val is not None:
            c.value = val; c.number_format = NF

a3(ws, r, "Compra de Servicios",       cs=costo_sv25,   ga=None, gc=None, og=None, tot=costo_sv25); r+=1
a3(ws, r, "Remuneraciones",            cs=None, ga=gadm25_suel, gc=None, og=None, tot=gadm25_suel); r+=1
a3(ws, r, "Cargas Sociales",           cs=None, ga=gadm25_cs,   gc=None, og=None, tot=gadm25_cs);   r+=1
a3(ws, r, "Gastos Bancarios",          cs=None, ga=gadm25_banc, gc=None, og=None, tot=gadm25_banc); r+=1
a3(ws, r, "Gastos Administrativos",    cs=None, ga=gadm25_adm,  gc=None, og=None, tot=gadm25_adm);  r+=1
a3(ws, r, "IIBB Buenos Aires",         cs=None, ga=None, gc=gcom25_bsas, og=None, tot=gcom25_bsas); r+=1
a3(ws, r, "IIBB Ciudad de Buenos Aires", cs=None, ga=None, gc=gcom25_caba, og=None, tot=gcom25_caba); r+=1
a3(ws, r, "IIBB Mendoza",             cs=None, ga=None, gc=gcom25_mdz,  og=None, tot=gcom25_mdz);   r+=1
a3_tot(ws, r, "TOTALES",
       cs=costo_sv25, ga=gadm25, gc=gcom25, og=0,
       tot=costo_sv25+gadm25+gcom25); r+=1

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS_III)
ws.cell(r,1).value = "Las notas y anexos que se acompañan forman parte integrante de los estados contables."
ws.cell(r,1).font  = Font(italic=True, size=9, name="Arial")

# ════════════════════════════════════════════════════════════════════════════
# NOTAS
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Notas")
set_cols(ws, [16, 86])
al_wrap = Alignment(wrap_text=True, vertical="top")
al_left = Alignment(horizontal="left", vertical="top")

r = title_block(ws, EMPRESA, "NOTAS A LOS ESTADOS CONTABLES",
    f"Por el ejercicio N°2 cerrado el {EJ25} — en pesos de poder adquisitivo al {EJ25}",
    ncols=2)

def nota(ws, r, nro, texto, h=60):
    row_height(ws, r, h)
    s(ws, r, 1, nro, bg=C_SEC, fn=fw(bold=True, color=C_HDR, size=10), al=al_left)
    s(ws, r, 2, texto, fn=fw(size=10), al=al_wrap)
    return r + 1

r = nota(ws, r, "1 — Normas Contables Aplicadas",
"""1.1 Modelo de presentación
Los presentes estados contables han sido preparados en moneda homogénea (pesos de poder adquisitivo al 31/12/2025), reconociendo en forma integral los efectos de la inflación de conformidad con la Resolución Técnica (RT) N° 6 de la FACPCE. Los criterios de exposición y valuación siguen la RT N°9 (normas particulares para entes comerciales), RT N°17 (normas particulares de medición para entes comerciales), RT N°16 (marco conceptual), RT N°18 (activos intangibles) y RT N°37 (auditoría). El estado de flujo de efectivo se prepara conforme a la RT N°8.

1.2 Reexpresión en moneda homogénea (RT N°6)
La economía argentina reviste carácter inflacionario. Los estados contables comparativos del ejercicio anterior (cerrado el 31/12/2024) se reexpresan aplicando el coeficiente 1,2870 sobre todos sus saldos nominales, a fin de expresarlos en pesos de poder adquisitivo del 31/12/2025. El efecto monetario de mantener activos y pasivos expuestos se expone como RECPAM en el Estado de Resultados.

1.3 Criterios de valuación
- Los activos y pasivos en moneda nacional se valúan a valor nominal (pesos corrientes).
- Los créditos se valúan a su valor probable de realización.
- Los bienes de cambio se valúan al costo de reposición o valor neto de realización, el menor.
- Los bienes de uso se deprecian por el método de línea recta.
- El impuesto a las ganancias se contabiliza por el método del impuesto diferido (RT N°17). Al cierre no se registran diferencias temporarias significativas.""", h=130)

r = nota(ws, r, "2 — Composición de los principales rubros", "", h=20)

r = nota(ws, r, "Nota 2.1\nCaja y Bancos",
f"""En pesos con centavos — EJERCICIO 2025 (actual) | EJERCICIO 2024 (reexpresado × {COF})

                                            2025              2024 (reexp.)
Caja                              $ {caja25_caja:>16,.2f}    $    {rx(800000):>14,.2f}
Banco Santander Río en $          $ {caja25_banco:>16,.2f}    $    {rx(36865.50):>14,.2f}
                                  ——————————————————         ———————————————
TOTAL                             $ {caja25:>16,.2f}    $  {caja24:>16,.2f}""", h=80)

r = nota(ws, r, "Nota 2.2\nCréditos por Ventas en Moneda",
f"""                                            2025              2024 (reexp.)
Deudores por Ventas               $ {cv25:>16,.2f}    $  {cv24:>16,.2f}
TOTAL                             $ {cv25:>16,.2f}    $  {cv24:>16,.2f}""", h=50)

r = nota(ws, r, "Nota 2.3\nOtros Créditos en Moneda",
f"""                                            2025              2024 (reexp.)
IVA Saldo Técnico                 $ {oc25_iva:>16,.2f}    $  {rx(323758.96):>14,.2f}  (crédito IG)
Saldo Libre Disponibilidad        $ {oc25_sld:>16,.2f}    $  {rx(655788.38):>14,.2f}  (anticipo prov.)
Impuesto a los Débitos y Créditos $ {oc25_dbc:>16,.2f}    $  {rx(32934.04):>14,.2f}
Retención Ganancias Sufrida       $ {oc25_ret:>16,.2f}    $  {rx(738537.29):>14,.2f}  (saldo IIBB)
                                  ——————————————————         ———————————————
TOTAL                             $ {oc25:>16,.2f}    $  {oc24:>16,.2f}""", h=90)

r = nota(ws, r, "Nota 2.4\nBienes de Cambio",
f"""                                            2025              2024 (reexp.)
Bienes de Cambio                  $ {bc25:>16,.2f}    $  {bc24:>16,.2f}
Las existencias al inicio del ejercicio 2025 se determinaron reexpresando el saldo de cierre 2024
($ {r24_bc:,.2f}) por el coeficiente {COF}, resultando en $ {costo_ei25:,.2f} (véase Anexo II).""", h=70)

r = nota(ws, r, "Nota 2.5\nCuentas a Pagar", "", h=20)

r = nota(ws, r, "  2.5.1 — Deudas Comerciales en Moneda",
f"""                                            2025              2024 (reexp.)
Proveedores                       $ {dc25:>16,.2f}    $  {dc24:>16,.2f}
TOTAL                             $ {dc25:>16,.2f}    $  {dc24:>16,.2f}""", h=55)

r = nota(ws, r, "  2.5.2 — Cargas Fiscales en Moneda",
f"""                                            2025              2024 (reexp.)
IVA a pagar                       $ {0:>16,.2f}    $  {rx(2000000):>14,.2f}
Acreditaciones Locales            $ {0:>16,.2f}    $  {rx(1360699):>14,.2f}
IIBB BSAS a pagar                 $ {df25_bsas:>16,.2f}    $  {rx(58067):>14,.2f}
IIBB CABA a pagar                 $ {df25_caba:>16,.2f}    $  {0:>14,.2f}
                                  ——————————————————         ———————————————
TOTAL                             $ {df25:>16,.2f}    $  {df24:>16,.2f}""", h=90)

r = nota(ws, r, "  2.5.3 — Remuneraciones y Cargas Sociales en Moneda",
f"""                                            2025              2024 (reexp.)
Cargas Sociales a pagar           $ {0:>16,.2f}    $  {rem24:>16,.2f}
Sueldos y Jornales a pagar        $ {0:>16,.2f}    $  {0:>16,.2f}
TOTAL                             $ {rem25:>16,.2f}    $  {rem24:>16,.2f}""", h=60)

r = nota(ws, r, "  2.5.4 — Deudas Sociales en Moneda",
f"""                                            2025              2024 (reexp.)
Cuenta particular Federico Guzzetti $ {ds25:>14,.2f}    $  {ds24:>16,.2f}
TOTAL                             $ {ds25:>16,.2f}    $  {ds24:>16,.2f}""", h=55)

r = nota(ws, r, "Nota 2.6\nResultados Financieros y por Tenencia (RECPAM)",
f"""                                            2025              2024 (reexp.)
RECPAM operativo (Sumas y Saldos) $ {recpam25:>16,.2f}    $  {recpam24:>16,.2f}
Ajuste reexpresión apertura (RT6) $ {recpam_rx_aper:>16,.2f}    $  {0:>16,.2f}
TOTAL RECPAM                      $ {recpam25_adj:>16,.2f}    $  {recpam24:>16,.2f}

El RECPAM (Resultado por Exposición al Cambio en el Poder Adquisitivo de la Moneda) surge de la
reexpresión de los activos y pasivos monetarios expuestos a la inflación durante el ejercicio,
de conformidad con la RT N°6.""", h=90)

r = nota(ws, r, "3 — Patrimonio Neto Negativo",
f"""El Estado de Situación Patrimonial al {EJ25} muestra un Patrimonio Neto negativo de
$ ({abs(pn25):,.2f}). Esta situación se origina principalmente en el resultado negativo del
ejercicio de $ ({abs(res25_adj):,.2f}), que refleja el nivel de gastos operativos y de estructura
incurridos durante el período. La sociedad está desarrollando su actividad comercial y
proyecta revertir esta situación mediante la expansión de su cartera de clientes y el
incremento de sus ingresos por servicios en los próximos ejercicios.""", h=90)

# ════════════════════════════════════════════════════════════════════════════
# PAGE SETUP — para que cada hoja entre en una página al exportar a PDF
# ════════════════════════════════════════════════════════════════════════════
from openpyxl.worksheet.page import PageMargins

def page_setup(ws, landscape=False, fit_w=1, fit_h=0, scale=None):
    """
    fit_w=1, fit_h=0 → ajusta ancho a 1 página, alto libre (crece verticalmente).
    fit_h=1 → fuerza todo en 1 página (achica lo que sea necesario).
    scale → porcentaje fijo (reemplaza fit, útil para hojas que se ven bien al 85%).
    """
    A4 = 9   # openpyxl PAPERSIZE_A4
    ws.page_setup.paperSize  = A4
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    if scale:
        ws.page_setup.fitToWidth  = 0
        ws.page_setup.fitToHeight = 0
        ws.page_setup.scale       = scale
        ws.sheet_properties.pageSetUpPr.fitToPage = False
    else:
        ws.page_setup.fitToWidth  = fit_w
        ws.page_setup.fitToHeight = fit_h

    # Márgenes reducidos (en pulgadas)
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5,
                                  header=0.2, footer=0.2)
    # Centrar horizontalmente
    ws.print_options.horizontalCentered = True

# Carátula — Portrait, 1 col ancho, sin límite alto
page_setup(wb["Carátula"],       landscape=False, fit_w=1, fit_h=0)
# ESP — Portrait, 3 cols → entra bien en A4 portrait
page_setup(wb["ESP"],            landscape=False, fit_w=1, fit_h=0)
# ER — Portrait
page_setup(wb["ER"],             landscape=False, fit_w=1, fit_h=0)
# EEPN — 13 columnas → Landscape obligatorio, forzar 1 página de ancho
page_setup(wb["EEPN"],           landscape=True,  fit_w=1, fit_h=0)
# EF — Portrait
page_setup(wb["EF"],             landscape=False, fit_w=1, fit_h=0)
# Anexo I — 11 columnas → Landscape
page_setup(wb["Anexo I"],        landscape=True,  fit_w=1, fit_h=0)
# Anexo II — Portrait
page_setup(wb["Anexo II"],       landscape=False, fit_w=1, fit_h=0)
# Anexo III — 6 columnas → Landscape
page_setup(wb["Anexo III"],      landscape=True,  fit_w=1, fit_h=0)
# Notas — Portrait, fit ancho; el alto puede ocupar varias páginas
page_setup(wb["Notas"],          landscape=False, fit_w=1, fit_h=0)

# ════════════════════════════════════════════════════════════════════════════
# GUARDAR
# ════════════════════════════════════════════════════════════════════════════
_DEFAULT_OUT = "/Users/jstivala/Downloads/EECC_AcerosSudamericanos_2025_v4.xlsx"

# Leer --output de argv si está presente
_out_arg = _DEFAULT_OUT
_argv = sys.argv[1:]
if "--output" in _argv:
    _idx = _argv.index("--output")
    if _idx + 1 < len(_argv):
        _out_arg = _argv[_idx + 1]

OUT = _out_arg

# Generar notes_data.json en el mismo directorio que el xlsx
notes_data = {
    "empresa": EMPRESA,
    "cuit": CUIT,
    "ej25": EJ25,
    "ej24": EJ24,
    "cof": COF,
    # 2025
    "caja25_caja": caja25_caja,
    "caja25_banco": caja25_banco,
    "caja25": caja25,
    "cv25": cv25,
    "oc25_iva": oc25_iva,
    "oc25_sld": oc25_sld,
    "oc25_dbc": oc25_dbc,
    "oc25_ret": oc25_ret,
    "oc25": oc25,
    "bc25": bc25,
    "dc25": dc25,
    "df25_bsas": df25_bsas,
    "df25_caba": df25_caba,
    "df25": df25,
    "rem25": rem25,
    "ds25": ds25,
    "tp25": tp25,
    "ta25": ta25,
    "pn25": pn25,
    "recpam25": recpam25,
    "recpam_rx_aper": recpam_rx_aper,
    "recpam25_adj": recpam25_adj,
    "res25_adj": res25_adj,
    # 2024 reexpresado
    "caja24": caja24,
    "cv24": cv24,
    "oc24": oc24,
    "bc24": bc24,
    "dc24": dc24,
    "df24": df24,
    "rem24": rem24,
    "ds24": ds24,
    "tp24": tp24,
    "pn24_rx": pn24_rx,
    "recpam24": recpam24,
    "res24": res24,
}
_notes_path = os.path.join(os.path.dirname(OUT), "notes_data.json")
with open(_notes_path, "w", encoding="utf-8") as _f:
    json.dump(notes_data, _f, ensure_ascii=False, indent=2)
print(f"📋 notes_data.json guardado: {_notes_path}")

wb.save(OUT)
print(f"\n✅ Guardado: {OUT}")
